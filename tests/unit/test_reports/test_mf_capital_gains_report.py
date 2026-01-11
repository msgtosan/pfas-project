"""Tests for MF Capital Gains Report Generator."""

import pytest
from datetime import date
from decimal import Decimal
from pathlib import Path
import tempfile

from pfas.reports.mf_capital_gains_report import (
    MFCapitalGainsReport,
    QuarterlySummary,
    FYSummary,
    TransactionDetail,
)
from pfas.parsers.mf.models import AssetClass


class TestQuarterlySummary:
    """Tests for QuarterlySummary dataclass."""

    def test_total_stcg_calculation(self):
        """Test total STCG is sum of equity and debt STCG."""
        summary = QuarterlySummary(
            quarter="Q1",
            start_date=date(2024, 4, 1),
            end_date=date(2024, 6, 30),
            equity_stcg=Decimal("10000"),
            debt_stcg=Decimal("5000"),
        )

        assert summary.total_stcg == Decimal("15000")

    def test_total_ltcg_calculation(self):
        """Test total LTCG is sum of equity and debt LTCG."""
        summary = QuarterlySummary(
            quarter="Q1",
            start_date=date(2024, 4, 1),
            end_date=date(2024, 6, 30),
            equity_ltcg=Decimal("200000"),
            debt_ltcg=Decimal("50000"),
        )

        assert summary.total_ltcg == Decimal("250000")

    def test_empty_quarter(self):
        """Test quarter with no transactions."""
        summary = QuarterlySummary(
            quarter="Q2",
            start_date=date(2024, 7, 1),
            end_date=date(2024, 9, 30),
        )

        assert summary.equity_stcg == Decimal("0")
        assert summary.equity_ltcg == Decimal("0")
        assert summary.debt_stcg == Decimal("0")
        assert summary.debt_ltcg == Decimal("0")
        assert summary.total_stcg == Decimal("0")
        assert summary.total_ltcg == Decimal("0")
        assert len(summary.equity_transactions) == 0
        assert len(summary.debt_transactions) == 0


class TestFYSummary:
    """Tests for FYSummary dataclass."""

    def test_total_stcg(self):
        """Test total STCG is sum of equity and debt."""
        summary = FYSummary(
            financial_year="2024-25",
            quarters=[],
            equity_stcg_total=Decimal("50000"),
            debt_stcg_total=Decimal("25000"),
        )

        assert summary.total_stcg == Decimal("75000")

    def test_total_ltcg(self):
        """Test total LTCG is sum of equity and debt."""
        summary = FYSummary(
            financial_year="2024-25",
            quarters=[],
            equity_ltcg_total=Decimal("300000"),
            debt_ltcg_total=Decimal("100000"),
        )

        assert summary.total_ltcg == Decimal("400000")

    def test_default_tax_rates(self):
        """Test default tax rates are set correctly."""
        summary = FYSummary(financial_year="2024-25", quarters=[])

        assert summary.equity_stcg_rate == Decimal("20")
        assert summary.equity_ltcg_rate == Decimal("12.5")


class TestMFCapitalGainsReport:
    """Tests for MFCapitalGainsReport class."""

    def test_initialization(self, db_connection):
        """Test report generator can be initialized."""
        report = MFCapitalGainsReport(db_connection)
        assert report.conn is not None

    def test_ltcg_exemption_constant(self, db_connection):
        """Test LTCG exemption is set to Rs 1.25 lakh."""
        report = MFCapitalGainsReport(db_connection)
        assert report.LTCG_EXEMPTION == Decimal("125000")

    def test_quarters_definition(self, db_connection):
        """Test quarters are defined correctly."""
        report = MFCapitalGainsReport(db_connection)

        assert len(report.QUARTERS) == 4
        assert report.QUARTERS[0][0] == "Q1"  # Apr-Jun
        assert report.QUARTERS[1][0] == "Q2"  # Jul-Sep
        assert report.QUARTERS[2][0] == "Q3"  # Oct-Dec
        assert report.QUARTERS[3][0] == "Q4"  # Jan-Mar

    def test_generate_empty_data(self, db_connection, sample_user):
        """Test generate with no transactions returns empty summary."""
        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(sample_user["id"], "2024-25")

        assert summary.financial_year == "2024-25"
        assert len(summary.quarters) == 4
        assert summary.equity_stcg_total == Decimal("0")
        assert summary.equity_ltcg_total == Decimal("0")
        assert summary.debt_stcg_total == Decimal("0")
        assert summary.debt_ltcg_total == Decimal("0")

    def test_quarter_date_ranges_fy_2024_25(self, db_connection, sample_user):
        """Test quarter date ranges are calculated correctly."""
        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(sample_user["id"], "2024-25")

        # Q1: Apr-Jun 2024
        assert summary.quarters[0].quarter == "Q1"
        assert summary.quarters[0].start_date == date(2024, 4, 1)
        assert summary.quarters[0].end_date == date(2024, 6, 30)

        # Q2: Jul-Sep 2024
        assert summary.quarters[1].quarter == "Q2"
        assert summary.quarters[1].start_date == date(2024, 7, 1)
        assert summary.quarters[1].end_date == date(2024, 9, 30)

        # Q3: Oct-Dec 2024
        assert summary.quarters[2].quarter == "Q3"
        assert summary.quarters[2].start_date == date(2024, 10, 1)
        assert summary.quarters[2].end_date == date(2024, 12, 31)

        # Q4: Jan-Mar 2025
        assert summary.quarters[3].quarter == "Q4"
        assert summary.quarters[3].start_date == date(2025, 1, 1)
        assert summary.quarters[3].end_date == date(2025, 3, 31)


class TestMFCapitalGainsReportWithData:
    """Tests for MFCapitalGainsReport with sample data."""

    @pytest.fixture
    def sample_mf_data(self, db_connection, sample_user):
        """Set up sample MF data for testing."""
        user_id = sample_user["id"]

        # Create AMC
        db_connection.execute(
            "INSERT INTO mf_amcs (name, short_name) VALUES (?, ?)",
            ("Test AMC", "TAMC"),
        )
        amc_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create Equity Scheme
        db_connection.execute(
            """
            INSERT INTO mf_schemes (amc_id, name, isin, asset_class, user_id)
            VALUES (?, ?, ?, ?, ?)
            """,
            (amc_id, "Test Equity Fund", "INE111111111", "EQUITY", user_id),
        )
        equity_scheme_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create Debt Scheme
        db_connection.execute(
            """
            INSERT INTO mf_schemes (amc_id, name, isin, asset_class, user_id)
            VALUES (?, ?, ?, ?, ?)
            """,
            (amc_id, "Test Debt Fund", "INE222222222", "DEBT", user_id),
        )
        debt_scheme_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create Folio for Equity
        db_connection.execute(
            """
            INSERT INTO mf_folios (user_id, scheme_id, folio_number, status)
            VALUES (?, ?, ?, ?)
            """,
            (user_id, equity_scheme_id, "1234567890", "ACTIVE"),
        )
        equity_folio_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Create Folio for Debt
        db_connection.execute(
            """
            INSERT INTO mf_folios (user_id, scheme_id, folio_number, status)
            VALUES (?, ?, ?, ?)
            """,
            (user_id, debt_scheme_id, "9876543210", "ACTIVE"),
        )
        debt_folio_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]

        db_connection.commit()

        return {
            "user_id": user_id,
            "equity_folio_id": equity_folio_id,
            "debt_folio_id": debt_folio_id,
        }

    def test_generate_with_equity_stcg(self, db_connection, sample_mf_data):
        """Test generate with equity STCG transaction."""
        user_id = sample_mf_data["user_id"]
        folio_id = sample_mf_data["equity_folio_id"]

        # Insert redemption transaction in Q1 (May 2024)
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                folio_id,
                "REDEMPTION",
                "2024-05-15",
                100,
                150.0,
                15000,
                "2024-01-15",
                10000,
                120,  # 4 months
                False,  # STCG
                5000,
                0,
                user_id,
            ),
        )
        db_connection.commit()

        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(user_id, "2024-25")

        assert summary.equity_stcg_total == Decimal("5000")
        assert summary.equity_ltcg_total == Decimal("0")
        assert summary.quarters[0].equity_stcg == Decimal("5000")
        assert len(summary.quarters[0].equity_transactions) == 1

    def test_generate_with_equity_ltcg(self, db_connection, sample_mf_data):
        """Test generate with equity LTCG transaction."""
        user_id = sample_mf_data["user_id"]
        folio_id = sample_mf_data["equity_folio_id"]

        # Insert redemption transaction in Q2 (Aug 2024) with LTCG
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                folio_id,
                "REDEMPTION",
                "2024-08-20",
                500,
                200.0,
                100000,
                "2022-06-20",
                50000,
                792,  # ~2.2 years
                True,  # LTCG
                0,
                50000,
                user_id,
            ),
        )
        db_connection.commit()

        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(user_id, "2024-25")

        assert summary.equity_stcg_total == Decimal("0")
        assert summary.equity_ltcg_total == Decimal("50000")
        assert summary.quarters[1].equity_ltcg == Decimal("50000")
        assert len(summary.quarters[1].equity_transactions) == 1

    def test_generate_with_debt_transaction(self, db_connection, sample_mf_data):
        """Test generate with debt fund transaction."""
        user_id = sample_mf_data["user_id"]
        folio_id = sample_mf_data["debt_folio_id"]

        # Insert debt redemption in Q3 (Nov 2024)
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                folio_id,
                "REDEMPTION",
                "2024-11-10",
                200,
                50.0,
                10000,
                "2024-05-10",
                8000,
                184,  # 6 months
                False,  # STCG
                2000,
                0,
                user_id,
            ),
        )
        db_connection.commit()

        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(user_id, "2024-25")

        assert summary.debt_stcg_total == Decimal("2000")
        assert summary.debt_ltcg_total == Decimal("0")
        assert summary.quarters[2].debt_stcg == Decimal("2000")
        assert len(summary.quarters[2].debt_transactions) == 1

    def test_ltcg_exemption_applied(self, db_connection, sample_mf_data):
        """Test LTCG exemption is correctly applied."""
        user_id = sample_mf_data["user_id"]
        folio_id = sample_mf_data["equity_folio_id"]

        # Insert large LTCG transaction (above exemption limit)
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                folio_id,
                "REDEMPTION",
                "2024-06-15",
                1000,
                500.0,
                500000,
                "2022-01-15",
                300000,
                882,
                True,
                0,
                200000,  # Rs 2 lakh LTCG
                user_id,
            ),
        )
        db_connection.commit()

        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(user_id, "2024-25")

        assert summary.equity_ltcg_total == Decimal("200000")
        assert summary.equity_ltcg_exemption == Decimal("125000")  # Rs 1.25L exemption
        assert summary.equity_taxable_ltcg == Decimal("75000")  # 2L - 1.25L

    def test_ltcg_below_exemption(self, db_connection, sample_mf_data):
        """Test LTCG below exemption limit results in zero taxable LTCG."""
        user_id = sample_mf_data["user_id"]
        folio_id = sample_mf_data["equity_folio_id"]

        # Insert small LTCG transaction (below exemption limit)
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                folio_id,
                "REDEMPTION",
                "2024-06-15",
                100,
                150.0,
                15000,
                "2022-01-15",
                10000,
                882,
                True,
                0,
                5000,  # Rs 5000 LTCG - below exemption
                user_id,
            ),
        )
        db_connection.commit()

        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(user_id, "2024-25")

        assert summary.equity_ltcg_total == Decimal("5000")
        assert summary.equity_ltcg_exemption == Decimal("5000")
        assert summary.equity_taxable_ltcg == Decimal("0")

    def test_multiple_quarters(self, db_connection, sample_mf_data):
        """Test transactions across multiple quarters."""
        user_id = sample_mf_data["user_id"]
        folio_id = sample_mf_data["equity_folio_id"]

        # Q1 transaction
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (folio_id, "REDEMPTION", "2024-05-15", 100, 150.0, 15000,
             "2024-01-15", 10000, 120, False, 5000, 0, user_id),
        )

        # Q3 transaction
        db_connection.execute(
            """
            INSERT INTO mf_transactions
            (folio_id, transaction_type, date, units, nav, amount,
             purchase_date, purchase_amount, holding_period_days,
             is_long_term, short_term_gain, long_term_gain, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (folio_id, "REDEMPTION", "2024-11-20", 200, 180.0, 36000,
             "2023-06-20", 28000, 518, True, 0, 8000, user_id),
        )

        db_connection.commit()

        report = MFCapitalGainsReport(db_connection)
        summary = report.generate(user_id, "2024-25")

        # Check Q1
        assert summary.quarters[0].equity_stcg == Decimal("5000")
        assert summary.quarters[0].equity_ltcg == Decimal("0")

        # Check Q2 (empty)
        assert summary.quarters[1].equity_stcg == Decimal("0")
        assert summary.quarters[1].equity_ltcg == Decimal("0")

        # Check Q3
        assert summary.quarters[2].equity_stcg == Decimal("0")
        assert summary.quarters[2].equity_ltcg == Decimal("8000")

        # Check totals
        assert summary.equity_stcg_total == Decimal("5000")
        assert summary.equity_ltcg_total == Decimal("8000")


class TestMFCapitalGainsReportExcel:
    """Tests for Excel export functionality."""

    @pytest.fixture
    def sample_summary(self):
        """Create a sample FYSummary for export testing."""
        quarters = []
        for i, (q_name, start_m, start_d, end_m, end_d) in enumerate(
            [("Q1", 4, 1, 6, 30), ("Q2", 7, 1, 9, 30), ("Q3", 10, 1, 12, 31), ("Q4", 1, 1, 3, 31)]
        ):
            year = 2024 if start_m >= 4 else 2025
            q = QuarterlySummary(
                quarter=q_name,
                start_date=date(year, start_m, start_d),
                end_date=date(year, end_m, end_d),
            )
            if q_name == "Q1":
                q.equity_stcg = Decimal("5000")
                q.equity_transactions.append(
                    TransactionDetail(
                        folio_number="12345",
                        scheme_name="Test Fund",
                        asset_class="EQUITY",
                        redemption_date=date(2024, 5, 15),
                        units=Decimal("100"),
                        nav=Decimal("150"),
                        amount=Decimal("15000"),
                        purchase_date=date(2024, 1, 15),
                        purchase_amount=Decimal("10000"),
                        holding_days=120,
                        is_long_term=False,
                        stcg=Decimal("5000"),
                        ltcg=Decimal("0"),
                    )
                )
            quarters.append(q)

        return FYSummary(
            financial_year="2024-25",
            quarters=quarters,
            equity_stcg_total=Decimal("5000"),
            equity_ltcg_total=Decimal("0"),
            equity_ltcg_exemption=Decimal("0"),
            equity_taxable_ltcg=Decimal("0"),
            debt_stcg_total=Decimal("0"),
            debt_ltcg_total=Decimal("0"),
        )

    def test_export_excel_creates_file(self, db_connection, sample_summary):
        """Test Excel export creates a file."""
        report = MFCapitalGainsReport(db_connection)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            result = report.export_excel(sample_summary, output_path)

            assert result.exists()
            assert result.suffix == ".xlsx"

    def test_export_excel_has_sheets(self, db_connection, sample_summary):
        """Test Excel export has correct sheets."""
        from openpyxl import load_workbook

        report = MFCapitalGainsReport(db_connection)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            report.export_excel(sample_summary, output_path)

            wb = load_workbook(output_path)

            assert "Summary" in wb.sheetnames
            assert "Q1" in wb.sheetnames
            assert "Q2" in wb.sheetnames
            assert "Q3" in wb.sheetnames
            assert "Q4" in wb.sheetnames

    def test_export_excel_summary_content(self, db_connection, sample_summary):
        """Test Excel summary sheet has correct content."""
        from openpyxl import load_workbook

        report = MFCapitalGainsReport(db_connection)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.xlsx"
            report.export_excel(sample_summary, output_path)

            wb = load_workbook(output_path)
            ws = wb["Summary"]

            # Check title
            assert "2024-25" in ws.cell(row=1, column=1).value


class TestMFCapitalGainsReportPDF:
    """Tests for PDF export functionality."""

    @pytest.fixture
    def sample_summary(self):
        """Create a sample FYSummary for export testing."""
        quarters = []
        for i, (q_name, start_m, start_d, end_m, end_d) in enumerate(
            [("Q1", 4, 1, 6, 30), ("Q2", 7, 1, 9, 30), ("Q3", 10, 1, 12, 31), ("Q4", 1, 1, 3, 31)]
        ):
            year = 2024 if start_m >= 4 else 2025
            q = QuarterlySummary(
                quarter=q_name,
                start_date=date(year, start_m, start_d),
                end_date=date(year, end_m, end_d),
            )
            quarters.append(q)

        return FYSummary(
            financial_year="2024-25",
            quarters=quarters,
            equity_stcg_total=Decimal("5000"),
            equity_ltcg_total=Decimal("150000"),
            equity_ltcg_exemption=Decimal("125000"),
            equity_taxable_ltcg=Decimal("25000"),
        )

    def test_export_pdf_with_reportlab(self, db_connection, sample_summary):
        """Test PDF export creates a file when reportlab is available."""
        report = MFCapitalGainsReport(db_connection)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.pdf"
            result = report.export_pdf(sample_summary, output_path)

            # Result is None if reportlab not installed, otherwise Path
            if result is not None:
                assert result.exists()
                assert result.suffix == ".pdf"

    def test_export_pdf_returns_none_without_reportlab(self, db_connection, sample_summary, monkeypatch):
        """Test PDF export returns None when reportlab is not available."""
        import builtins

        original_import = builtins.__import__

        def mock_import(name, *args, **kwargs):
            if name.startswith("reportlab"):
                raise ImportError("Mock reportlab import error")
            return original_import(name, *args, **kwargs)

        # Create fresh report instance to test import handling
        report = MFCapitalGainsReport(db_connection)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_report.pdf"

            # Patch import inside the method
            monkeypatch.setattr(builtins, "__import__", mock_import)

            result = report.export_pdf(sample_summary, output_path)
            assert result is None
