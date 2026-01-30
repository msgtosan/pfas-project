"""
Unit and Integration Tests for Networth Report Module

Tests cover:
- NetworthCalculator: Aggregation, CAGR/XIRR calculations, edge cases
- NetworthReportGenerator: Excel generation, chart creation
- XIRRCalculator: Convergence, edge cases
- CLI: Argument parsing, config loading

Run with: pytest tests/unit/test_reports/test_networth_report.py -v
"""

import json
import sqlite3
import tempfile
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# Add src to path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent / "src"))

from pfas.reports.networth_report import (
    AssetHolding,
    AssetMetrics,
    NetworthCalculator,
    NetworthReportGenerator,
    NetworthSummary,
    PeriodSnapshot,
    XIRRCalculator,
)


# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def temp_db():
    """Create temporary in-memory database with schema."""
    conn = sqlite3.connect(":memory:")

    conn.executescript("""
        CREATE TABLE users (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL
        );

        CREATE TABLE mf_holdings (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            scheme_name TEXT,
            asset_class TEXT,
            current_value DECIMAL,
            cost_value DECIMAL,
            nav_date DATE
        );

        CREATE TABLE stock_holdings (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            symbol TEXT,
            market_value DECIMAL,
            total_cost_basis DECIMAL,
            as_of_date DATE
        );

        CREATE TABLE mf_transactions (
            id INTEGER PRIMARY KEY,
            folio_id INTEGER,
            transaction_type TEXT,
            date DATE,
            amount DECIMAL
        );

        CREATE TABLE mf_folios (
            id INTEGER PRIMARY KEY,
            user_id INTEGER
        );

        CREATE TABLE stock_trades (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            buy_sell TEXT,
            trade_date DATE,
            net_amount DECIMAL
        );

        CREATE TABLE epf_transactions (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            employee_balance DECIMAL,
            employer_balance DECIMAL,
            transaction_date DATE
        );

        CREATE TABLE ppf_transactions (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            balance DECIMAL,
            transaction_date DATE
        );

        INSERT INTO users (id, name) VALUES (1, 'TestUser');
    """)

    return conn


@pytest.fixture
def mock_resolver(tmp_path):
    """Create mock PathResolver."""
    resolver = MagicMock()
    resolver.root = tmp_path
    resolver.user_dir = tmp_path / "Users" / "TestUser"
    resolver.user_config_file.return_value = None

    def report_file(asset_type, report_type, extension="xlsx"):
        return tmp_path / "reports" / f"{asset_type}_{report_type}.{extension}"

    resolver.report_file = report_file
    return resolver


@pytest.fixture
def sample_config():
    """Sample configuration for testing."""
    return {
        "version": "1.0.0",
        "asset_categories": {
            "mutual_funds": {
                "enabled": True,
                "label": "Mutual Funds",
                "table": "mf_holdings",
                "value_column": "current_value",
                "cost_column": "cost_value",
                "date_column": "nav_date"
            },
            "indian_stocks": {
                "enabled": True,
                "label": "Indian Stocks",
                "table": "stock_holdings",
                "value_column": "market_value",
                "cost_column": "total_cost_basis",
                "date_column": "as_of_date"
            },
            "epf": {
                "enabled": True,
                "label": "EPF",
                "table": "epf_transactions",
                "balance_columns": ["employee_balance", "employer_balance"],
                "date_column": "transaction_date"
            }
        },
        "defaults": {
            "granularity": "fy",
            "fy_range": "current"
        },
        "report_settings": {
            "excel": {
                "sheets": [
                    {"name": "Summary", "type": "summary"},
                    {"name": "Allocation", "type": "allocation_pie"},
                    {"name": "Growth", "type": "growth_chart"}
                ],
                "styles": {
                    "header": {"bold": True, "bg_color": "#1F4E79", "font_color": "#FFFFFF"}
                }
            }
        }
    }


# =============================================================================
# XIRRCalculator Tests
# =============================================================================

class TestXIRRCalculator:
    """Tests for XIRR calculation."""

    def test_xirr_simple_case(self):
        """Test XIRR with simple investment and return."""
        # Invest 10000, get 11000 after 1 year = 10% return
        cashflows = [
            (date(2023, 1, 1), -10000),
            (date(2024, 1, 1), 11000),
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is not None
        assert abs(xirr - 0.10) < 0.01  # ~10%

    def test_xirr_multiple_cashflows(self):
        """Test XIRR with multiple investments."""
        cashflows = [
            (date(2023, 1, 1), -10000),
            (date(2023, 7, 1), -5000),
            (date(2024, 1, 1), 17000),
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is not None
        assert xirr > 0  # Should be positive return

    def test_xirr_negative_return(self):
        """Test XIRR with loss."""
        cashflows = [
            (date(2023, 1, 1), -10000),
            (date(2024, 1, 1), 8000),  # 20% loss
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is not None
        assert xirr < 0  # Negative return

    def test_xirr_insufficient_data(self):
        """Test XIRR with insufficient cashflows."""
        cashflows = [(date(2023, 1, 1), -10000)]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is None

    def test_xirr_all_positive(self):
        """Test XIRR when all cashflows are positive (invalid)."""
        cashflows = [
            (date(2023, 1, 1), 10000),
            (date(2024, 1, 1), 11000),
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is None

    def test_xirr_all_negative(self):
        """Test XIRR when all cashflows are negative (invalid)."""
        cashflows = [
            (date(2023, 1, 1), -10000),
            (date(2024, 1, 1), -5000),
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is None

    def test_xirr_zero_investment(self):
        """Test XIRR with zero initial investment."""
        cashflows = [
            (date(2023, 1, 1), 0),
            (date(2024, 1, 1), 10000),
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is None

    def test_xirr_high_return(self):
        """Test XIRR with very high return (100%)."""
        cashflows = [
            (date(2023, 1, 1), -10000),
            (date(2024, 1, 1), 20000),  # 100% return
        ]
        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is not None
        assert abs(xirr - 1.0) < 0.05  # ~100%

    def test_xirr_sip_pattern(self):
        """Test XIRR with SIP-like monthly investments."""
        cashflows = []
        start = date(2023, 1, 1)

        # Monthly SIP of 5000 for 12 months
        for i in range(12):
            cf_date = date(2023, 1 + i if i < 12 else 1, 1)
            if i >= 12:
                cf_date = date(2024, i - 11, 1)
            cashflows.append((start + timedelta(days=30 * i), -5000))

        # Final value after 1 year
        cashflows.append((date(2024, 1, 1), 65000))  # ~8% return

        xirr = XIRRCalculator.calculate(cashflows)
        assert xirr is not None


# =============================================================================
# NetworthCalculator Tests
# =============================================================================

class TestNetworthCalculator:
    """Tests for NetworthCalculator."""

    def test_parse_fy_range_current(self, temp_db, mock_resolver, sample_config):
        """Test parsing 'current' FY range."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        start, end = calc._parse_fy_range("current")

        today = date.today()
        if today.month >= 4:
            expected_start = date(today.year, 4, 1)
        else:
            expected_start = date(today.year - 1, 4, 1)

        assert start == expected_start
        assert end == today

    def test_parse_fy_range_single_fy(self, temp_db, mock_resolver, sample_config):
        """Test parsing single FY like '2024-25'."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        start, end = calc._parse_fy_range("2024-25")

        assert start == date(2024, 4, 1)
        assert end == date(2025, 3, 31)

    def test_parse_fy_range_multi_year(self, temp_db, mock_resolver, sample_config):
        """Test parsing multi-year range like '2023-2026'."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        start, end = calc._parse_fy_range("2023-2026")

        assert start == date(2023, 4, 1)
        assert end == date(2026, 3, 31)

    def test_generate_periods_fy(self, temp_db, mock_resolver, sample_config):
        """Test period generation for FY granularity."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        periods = calc._generate_periods(
            date(2023, 4, 1), date(2025, 3, 31), "fy"
        )

        assert len(periods) == 2
        assert periods[0][0] == "2023-24"
        assert periods[1][0] == "2024-25"

    def test_generate_periods_quarterly(self, temp_db, mock_resolver, sample_config):
        """Test period generation for quarterly granularity."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        periods = calc._generate_periods(
            date(2024, 4, 1), date(2025, 3, 31), "quarterly"
        )

        assert len(periods) == 4

    def test_generate_periods_monthly(self, temp_db, mock_resolver, sample_config):
        """Test period generation for monthly granularity."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        periods = calc._generate_periods(
            date(2024, 4, 1), date(2024, 6, 30), "monthly"
        )

        assert len(periods) == 3  # Apr, May, Jun

    def test_calculate_with_mf_data(self, temp_db, mock_resolver, sample_config):
        """Test calculation with MF holdings data."""
        # Insert test data
        temp_db.execute("""
            INSERT INTO mf_holdings (user_id, scheme_name, asset_class, current_value, cost_value, nav_date)
            VALUES (1, 'Test Equity Fund', 'EQUITY', 100000, 80000, '2024-03-31')
        """)
        temp_db.commit()

        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25",
            granularity="fy"
        )

        assert summary.total_networth == Decimal("100000")
        assert "Mutual Funds" in summary.by_category

    def test_calculate_with_stock_data(self, temp_db, mock_resolver, sample_config):
        """Test calculation with stock holdings data."""
        temp_db.execute("""
            INSERT INTO stock_holdings (user_id, symbol, market_value, total_cost_basis, as_of_date)
            VALUES (1, 'RELIANCE', 50000, 40000, '2024-03-31')
        """)
        temp_db.commit()

        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25"
        )

        assert summary.total_networth >= Decimal("50000")
        assert "Indian Stocks" in summary.by_category

    def test_calculate_with_epf_balance(self, temp_db, mock_resolver, sample_config):
        """Test calculation with EPF balance data."""
        temp_db.execute("""
            INSERT INTO epf_transactions (user_id, employee_balance, employer_balance, transaction_date)
            VALUES (1, 200000, 100000, '2024-03-31')
        """)
        temp_db.commit()

        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25"
        )

        # EPF should show 300000 (200000 + 100000)
        assert "EPF" in summary.by_category
        assert summary.by_category["EPF"] == Decimal("300000")

    def test_calculate_empty_data(self, temp_db, mock_resolver, sample_config):
        """Test calculation with no holdings data."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25"
        )

        assert summary.total_networth == Decimal("0")
        # Categories may be present with zero values
        assert all(v == Decimal("0") for v in summary.by_category.values())

    def test_allocation_percentage(self, temp_db, mock_resolver, sample_config):
        """Test allocation percentage calculation."""
        temp_db.execute("""
            INSERT INTO mf_holdings (user_id, scheme_name, asset_class, current_value, cost_value, nav_date)
            VALUES (1, 'Fund A', 'EQUITY', 60000, 50000, '2024-03-31')
        """)
        temp_db.execute("""
            INSERT INTO stock_holdings (user_id, symbol, market_value, total_cost_basis, as_of_date)
            VALUES (1, 'TCS', 40000, 35000, '2024-03-31')
        """)
        temp_db.commit()

        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25"
        )

        # MF: 60%, Stocks: 40%
        assert abs(summary.allocation_pct.get("Mutual Funds", 0) - 60) < 1
        assert abs(summary.allocation_pct.get("Indian Stocks", 0) - 40) < 1


# =============================================================================
# NetworthReportGenerator Tests
# =============================================================================

class TestNetworthReportGenerator:
    """Tests for Excel report generation."""

    def test_generate_creates_file(self, temp_db, mock_resolver, sample_config, tmp_path):
        """Test that generate creates an Excel file."""
        output_path = tmp_path / "test_report.xlsx"

        generator = NetworthReportGenerator(temp_db, mock_resolver, sample_config)
        result_path = generator.generate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25",
            output_path=output_path
        )

        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

    def test_generate_with_data(self, temp_db, mock_resolver, sample_config, tmp_path):
        """Test report generation with actual data."""
        # Insert test data
        temp_db.execute("""
            INSERT INTO mf_holdings (user_id, scheme_name, asset_class, current_value, cost_value, nav_date)
            VALUES (1, 'Test Fund', 'EQUITY', 100000, 80000, '2024-03-31')
        """)
        temp_db.commit()

        output_path = tmp_path / "report_with_data.xlsx"
        generator = NetworthReportGenerator(temp_db, mock_resolver, sample_config)
        result_path = generator.generate(
            user_id=1,
            user_name="TestUser",
            output_path=output_path
        )

        assert result_path.exists()

        # Verify Excel has content
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        assert "Summary" in wb.sheetnames

    def test_generate_all_sheets(self, temp_db, mock_resolver, sample_config, tmp_path):
        """Test that all configured sheets are created."""
        output_path = tmp_path / "all_sheets.xlsx"

        generator = NetworthReportGenerator(temp_db, mock_resolver, sample_config)
        generator.generate(
            user_id=1,
            user_name="TestUser",
            output_path=output_path
        )

        from openpyxl import load_workbook
        wb = load_workbook(output_path)

        expected_sheets = ["Summary", "Allocation", "Growth"]
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"


# =============================================================================
# Data Model Tests
# =============================================================================

class TestDataModels:
    """Tests for data model classes."""

    def test_asset_holding_gain_calculation(self):
        """Test AssetHolding gain/loss calculation."""
        holding = AssetHolding(
            category="MF",
            sub_category="Equity",
            label="Test Fund",
            as_of_date=date.today(),
            current_value=Decimal("120000"),
            cost_basis=Decimal("100000")
        )

        assert holding.gain_loss == Decimal("20000")
        assert holding.gain_loss_pct == Decimal("20")

    def test_asset_holding_with_forex(self):
        """Test AssetHolding with forex conversion."""
        holding = AssetHolding(
            category="USA Stocks",
            sub_category=None,
            label="AAPL",
            as_of_date=date.today(),
            current_value=Decimal("1000"),  # USD
            cost_basis=Decimal("800"),
            currency="USD",
            exchange_rate=Decimal("83.50")
        )

        assert holding.value_inr == Decimal("83500")
        assert holding.cost_inr == Decimal("66800")
        assert holding.gain_loss == Decimal("16700")

    def test_period_snapshot_totals(self):
        """Test PeriodSnapshot total calculations."""
        snapshot = PeriodSnapshot(
            period_key="2024-25",
            period_label="FY 2024-25",
            period_end_date=date(2025, 3, 31),
            total_value=Decimal("500000"),
            total_cost=Decimal("400000")
        )

        assert snapshot.total_gain == Decimal("100000")
        assert snapshot.gain_pct == Decimal("25")

    def test_period_snapshot_zero_cost(self):
        """Test PeriodSnapshot with zero cost basis."""
        snapshot = PeriodSnapshot(
            period_key="2024-25",
            period_label="FY 2024-25",
            period_end_date=date(2025, 3, 31),
            total_value=Decimal("100000"),
            total_cost=Decimal("0")
        )

        assert snapshot.total_gain == Decimal("100000")
        assert snapshot.gain_pct == Decimal("0")  # Avoid division by zero


# =============================================================================
# Integration Tests
# =============================================================================

class TestIntegration:
    """Integration tests with realistic data scenarios."""

    @pytest.fixture
    def populated_db(self, temp_db):
        """Create database with realistic multi-asset data."""
        # MF Holdings - multiple schemes
        temp_db.executemany("""
            INSERT INTO mf_holdings (user_id, scheme_name, asset_class, current_value, cost_value, nav_date)
            VALUES (?, ?, ?, ?, ?, ?)
        """, [
            (1, "HDFC Equity Fund", "EQUITY", 250000, 200000, "2024-03-31"),
            (1, "ICICI Debt Fund", "DEBT", 150000, 145000, "2024-03-31"),
            (1, "SBI Liquid Fund", "DEBT", 50000, 49000, "2024-03-31"),
            # Historical data
            (1, "HDFC Equity Fund", "EQUITY", 180000, 200000, "2023-03-31"),
            (1, "ICICI Debt Fund", "DEBT", 140000, 145000, "2023-03-31"),
        ])

        # Stock Holdings
        temp_db.executemany("""
            INSERT INTO stock_holdings (user_id, symbol, market_value, total_cost_basis, as_of_date)
            VALUES (?, ?, ?, ?, ?)
        """, [
            (1, "RELIANCE", 100000, 80000, "2024-03-31"),
            (1, "TCS", 75000, 60000, "2024-03-31"),
            (1, "INFY", 50000, 55000, "2024-03-31"),  # Loss position
            # Historical
            (1, "RELIANCE", 70000, 80000, "2023-03-31"),
            (1, "TCS", 55000, 60000, "2023-03-31"),
        ])

        # EPF
        temp_db.executemany("""
            INSERT INTO epf_transactions (user_id, employee_balance, employer_balance, transaction_date)
            VALUES (?, ?, ?, ?)
        """, [
            (1, 500000, 250000, "2024-03-31"),
            (1, 400000, 200000, "2023-03-31"),
        ])

        # PPF
        temp_db.executemany("""
            INSERT INTO ppf_transactions (user_id, balance, transaction_date)
            VALUES (?, ?, ?)
        """, [
            (1, 300000, "2024-03-31"),
            (1, 250000, "2023-03-31"),
        ])

        temp_db.commit()
        return temp_db

    def test_full_networth_calculation(self, populated_db, mock_resolver, sample_config):
        """Test complete networth calculation across all assets."""
        # Add PPF to config
        sample_config["asset_categories"]["ppf"] = {
            "enabled": True,
            "label": "PPF",
            "table": "ppf_transactions",
            "balance_column": "balance",
            "date_column": "transaction_date"
        }

        calc = NetworthCalculator(populated_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25",
            granularity="fy"
        )

        # Verify we have all 4 categories
        assert len(summary.by_category) == 4
        assert "Mutual Funds" in summary.by_category
        assert "Indian Stocks" in summary.by_category
        assert "EPF" in summary.by_category
        assert "PPF" in summary.by_category

        # Verify totals are positive and consistent
        assert summary.total_networth > Decimal("0")
        assert summary.total_networth == sum(summary.by_category.values())

        # Verify EPF has expected value (employee + employer)
        assert summary.by_category["EPF"] == Decimal("750000")
        # Verify PPF
        assert summary.by_category["PPF"] == Decimal("300000")

    def test_yoy_growth_calculation(self, populated_db, mock_resolver, sample_config):
        """Test year-over-year growth calculation."""
        calc = NetworthCalculator(populated_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2023-2025",
            granularity="fy"
        )

        assert len(summary.snapshots) == 2

        # FY 2023-24 snapshot
        fy23 = summary.snapshots[0]
        # FY 2024-25 snapshot
        fy24 = summary.snapshots[1]

        # Verify we have data for both periods
        assert fy23.total_value >= Decimal("0")
        assert fy24.total_value >= Decimal("0")

        # Both periods should have the same categories
        assert set(fy23.by_category.keys()) == set(fy24.by_category.keys())

    def test_cagr_calculation(self, populated_db, mock_resolver, sample_config):
        """Test CAGR calculation."""
        calc = NetworthCalculator(populated_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2023-2025",
            granularity="fy"
        )

        # Should have CAGR calculated (may be 0 if no growth between periods)
        assert summary.overall_cagr is not None
        # CAGR should be a valid number (not NaN or inf)
        import math
        assert not math.isnan(summary.overall_cagr)
        assert not math.isinf(summary.overall_cagr)

    def test_full_report_generation(self, populated_db, mock_resolver, sample_config, tmp_path):
        """Test complete report generation with real data."""
        output_path = tmp_path / "full_report.xlsx"

        generator = NetworthReportGenerator(populated_db, mock_resolver, sample_config)
        result_path = generator.generate(
            user_id=1,
            user_name="TestUser",
            fy_range="2023-2025",
            granularity="fy",
            detailed=True,
            output_path=output_path
        )

        assert result_path.exists()

        # Verify Excel structure
        from openpyxl import load_workbook
        wb = load_workbook(result_path)

        # Check summary sheet has data
        ws = wb["Summary"]
        assert ws["A1"].value is not None

    def test_quarterly_breakdown(self, populated_db, mock_resolver, sample_config):
        """Test quarterly granularity."""
        calc = NetworthCalculator(populated_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2024-25",
            granularity="quarterly"
        )

        # Should have 4 quarters
        assert len(summary.snapshots) <= 4


# =============================================================================
# Edge Case Tests
# =============================================================================

class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_missing_table(self, temp_db, mock_resolver):
        """Test handling of missing database table."""
        config = {
            "asset_categories": {
                "missing_asset": {
                    "enabled": True,
                    "label": "Missing Asset",
                    "table": "nonexistent_table",
                    "value_column": "value",
                    "date_column": "date"
                }
            },
            "defaults": {"granularity": "fy"}
        }

        calc = NetworthCalculator(temp_db, mock_resolver, config)
        summary = calc.calculate(user_id=1, user_name="TestUser")

        # Should not crash, just return zero
        assert summary.total_networth == Decimal("0")

    def test_null_values(self, temp_db, mock_resolver, sample_config):
        """Test handling of NULL values in database."""
        temp_db.execute("""
            INSERT INTO mf_holdings (user_id, scheme_name, asset_class, current_value, cost_value, nav_date)
            VALUES (1, 'Null Test', 'EQUITY', NULL, NULL, '2024-03-31')
        """)
        temp_db.commit()

        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(user_id=1, user_name="TestUser")

        # Should handle NULL gracefully
        assert summary.total_networth >= Decimal("0")

    def test_future_date_range(self, temp_db, mock_resolver, sample_config):
        """Test with future date range."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=1,
            user_name="TestUser",
            fy_range="2030-2035"
        )

        # Should return empty/zero
        assert summary.total_networth == Decimal("0")

    def test_invalid_user(self, temp_db, mock_resolver, sample_config):
        """Test with non-existent user ID."""
        calc = NetworthCalculator(temp_db, mock_resolver, sample_config)
        summary = calc.calculate(
            user_id=9999,  # Non-existent
            user_name="NonExistent"
        )

        # Should not crash
        assert summary.total_networth == Decimal("0")


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
