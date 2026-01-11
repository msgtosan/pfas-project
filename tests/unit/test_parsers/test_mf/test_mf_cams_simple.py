#!/usr/bin/env python3
"""Simple integration test for MF CAMS - reads Excel structure without pandas."""

import sys
from pathlib import Path

# File path
cams_file = Path.home() / "projects/pfas-project/Data/Users/Sanjay/Mutual-Fund/CAMS/Sanjay_CAMS_CG_FY2024-25_v1.xlsx"

print("="*70)
print("MF CAMS Integration Test - Real Data Analysis")
print("="*70)

print(f"\n📁 File Information:")
print(f"   Path: {cams_file}")
print(f"   Exists: {cams_file.exists()}")
if cams_file.exists():
    print(f"   Size: {cams_file.stat().st_size / 1024:.1f} KB")
    print(f"   Modified: {Path(cams_file).stat().st_mtime}")

# Try to read with openpyxl
try:
    from openpyxl import load_workbook
    print("\n✅ openpyxl available - reading file structure...\n")

    wb = load_workbook(cams_file)
    print(f"📊 Excel Sheet Structure:")
    print(f"   Sheets: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   Sheet: '{sheet_name}'")
        print(f"      Rows: {ws.max_row}")
        print(f"      Columns: {ws.max_column}")

        # Get headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))

        if headers:
            print(f"      Headers: {headers[:8]}")
            if len(headers) > 8:
                print(f"               ... and {len(headers) - 8} more columns")

        # Sample first 3 data rows
        print(f"\n      Sample Data (first 3 rows):")
        for row_idx in range(2, min(5, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(5, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                row_data.append(str(cell.value)[:30] if cell.value else "")
            print(f"         Row {row_idx}: {row_data}")

        print()

except ImportError:
    print("⚠️  openpyxl not available - reading via zipfile...\n")

    import zipfile
    import xml.etree.ElementTree as ET

    try:
        with zipfile.ZipFile(cams_file, 'r') as z:
            print(f"📦 Excel File Contents (ZIP structure):")
            print(f"   Files: {z.namelist()[:10]}")

            # Read workbook.xml to get sheet info
            if 'xl/workbook.xml' in z.namelist():
                workbook_xml = z.read('xl/workbook.xml')
                root = ET.fromstring(workbook_xml)

                # Find sheet names
                sheets = []
                for sheet in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                    sheets.append(sheet.get('name'))

                print(f"\n📊 Sheets found: {sheets}\n")

    except Exception as e:
        print(f"⚠️  Could not read file structure: {e}")

# Analyze file content
print("\n" + "="*70)
print("📋 File Content Summary:")
print("="*70)

try:
    from openpyxl import load_workbook

    wb = load_workbook(cams_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Count non-empty rows
        non_empty_rows = 0
        for row in ws.iter_rows():
            if any(cell.value for cell in row):
                non_empty_rows += 1

        print(f"\n📊 Sheet: {sheet_name}")
        print(f"   Non-empty rows: {non_empty_rows}")

        # Get first header row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))

        print(f"   Columns ({len(headers)}):")
        for i, header in enumerate(headers[:15], 1):
            print(f"      {i}. {header}")

        if len(headers) > 15:
            print(f"      ... and {len(headers) - 15} more columns")

        # Count rows by type
        print(f"\n   Data Analysis:")

        if "Scheme Name" in headers or "SCHEME NAME" in headers:
            scheme_col = next((i for i, h in enumerate(headers, 1) if "Scheme" in h), None)
            if scheme_col:
                schemes = set()
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row_idx, scheme_col)
                    if cell.value:
                        schemes.add(str(cell.value))
                print(f"      Unique Schemes: {len(schemes)}")

        if "Folio" in headers or "FOLIO" in headers or "Folio No" in headers:
            folio_col = next((i for i, h in enumerate(headers, 1) if "Folio" in h), None)
            if folio_col:
                folios = set()
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row_idx, folio_col)
                    if cell.value:
                        folios.add(str(cell.value))
                print(f"      Unique Folios: {len(folios)}")

except Exception as e:
    print(f"\n⚠️  Error analyzing content: {e}")

print("\n" + "="*70)
print("✅ File analysis complete")
print("="*70)
