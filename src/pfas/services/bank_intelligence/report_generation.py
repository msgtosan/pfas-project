"""
Fiscal Report Generator for Bank Intelligence.

Generates Excel Master Reports with:
- Detailed Ledger sheets with full transaction metadata
- FY Summary with category breakdown
- Auto-filters on all headers
- Dynamic SUBTOTAL formulas
- Conditional formatting
"""

import sqlite3
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
from typing import Optional, List, Dict, Any, Tuple
from collections import defaultdict

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class FiscalReportGenerator:
    """
    Generates fiscal year reports from bank intelligence database.

    Features:
    - Detailed ledger with uid, user_name, base_string
    - FY pivot by fiscal year and category
    - Category-wise breakdown
    - Auto-filters and SUBTOTAL formulas
    """

    def __init__(self, db_path: str):
        """
        Initialize generator.

        Args:
            db_path: Path to money_movement.db
        """
        self.db_path = db_path
        self.conn: Optional[sqlite3.Connection] = None

    def connect(self) -> None:
        """Connect to database."""
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row

    def close(self) -> None:
        """Close database connection."""
        if self.conn:
            self.conn.close()
            self.conn = None

    def __enter__(self) -> "FiscalReportGenerator":
        """Context manager entry."""
        self.connect()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """Context manager exit."""
        self.close()

    def generate_master_report(
        self,
        output_path: str,
        fiscal_year: Optional[str] = None,
        user_name: Optional[str] = None
    ) -> str:
        """
        Generate Master Report Excel file.

        Args:
            output_path: Path for output Excel file
            fiscal_year: Optional filter by fiscal year (e.g., "FY 2024-25")
            user_name: Optional filter by user

        Returns:
            Path to generated file
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required for Excel generation")

        if not self.conn:
            self.connect()

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Generate sheets
        self._create_detailed_ledger(wb, fiscal_year, user_name)
        self._create_fy_summary(wb, user_name)
        self._create_category_analysis(wb, fiscal_year, user_name)
        self._create_income_summary(wb, fiscal_year, user_name)

        # Save workbook
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path

    def _create_detailed_ledger(
        self,
        wb: "Workbook",
        fiscal_year: Optional[str],
        user_name: Optional[str]
    ) -> None:
        """Create Detailed Ledger sheet."""
        ws = wb.create_sheet("Detailed_Ledger")

        # Query data
        query = """
            SELECT uid, user_name, bank_name, txn_date, remarks, base_string,
                   amount, txn_type, balance, category, fiscal_year, source_file
            FROM bank_transactions_intel
            WHERE 1=1
        """
        params = []

        if fiscal_year:
            query += " AND fiscal_year = ?"
            params.append(fiscal_year)
        if user_name:
            query += " AND user_name = ?"
            params.append(user_name)

        query += " ORDER BY txn_date DESC, created_at DESC"

        cursor = self.conn.execute(query, params)
        rows = cursor.fetchall()

        # Headers
        headers = [
            "UID", "User", "Bank", "Date", "Remarks", "Base String",
            "Amount", "Type", "Balance", "Category", "Fiscal Year", "Source"
        ]

        self._write_headers(ws, headers)

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row["uid"][:12] + "...")
            ws.cell(row=row_idx, column=2, value=row["user_name"])
            ws.cell(row=row_idx, column=3, value=row["bank_name"])
            ws.cell(row=row_idx, column=4, value=row["txn_date"])
            ws.cell(row=row_idx, column=5, value=row["remarks"][:50] if row["remarks"] else "")
            ws.cell(row=row_idx, column=6, value=row["base_string"][:80] if row["base_string"] else "")

            # Amount with formatting
            amount = Decimal(row["amount"]) if row["amount"] else Decimal(0)
            cell = ws.cell(row=row_idx, column=7, value=float(amount))
            cell.number_format = '#,##0.00'

            # Color based on credit/debit
            if amount > 0:
                cell.font = Font(color="006600")  # Green for credit
            else:
                cell.font = Font(color="CC0000")  # Red for debit

            ws.cell(row=row_idx, column=8, value=row["txn_type"])

            if row["balance"]:
                balance_cell = ws.cell(row=row_idx, column=9, value=float(Decimal(row["balance"])))
                balance_cell.number_format = '#,##0.00'

            ws.cell(row=row_idx, column=10, value=row["category"])
            ws.cell(row=row_idx, column=11, value=row["fiscal_year"])
            ws.cell(row=row_idx, column=12, value=Path(row["source_file"]).name if row["source_file"] else "")

        # Add SUBTOTAL formula for amount column
        if len(rows) > 0:
            subtotal_row = len(rows) + 2
            ws.cell(row=subtotal_row, column=6, value="SUBTOTAL:")
            subtotal_cell = ws.cell(
                row=subtotal_row, column=7,
                value=f"=SUBTOTAL(9,G2:G{len(rows)+1})"
            )
            subtotal_cell.font = Font(bold=True)
            subtotal_cell.number_format = '#,##0.00'

        # Apply auto-filter
        ws.auto_filter.ref = f"A1:L{max(2, len(rows)+1)}"

        # Freeze header row
        ws.freeze_panes = "A2"

        # Adjust column widths
        self._adjust_column_widths(ws)

    def _create_fy_summary(
        self,
        wb: "Workbook",
        user_name: Optional[str]
    ) -> None:
        """Create FY Summary pivot sheet."""
        ws = wb.create_sheet("FY_Summary")

        # Query aggregated data
        query = """
            SELECT fiscal_year, category, txn_type,
                   COUNT(*) as txn_count,
                   SUM(CAST(amount AS REAL)) as total_amount
            FROM bank_transactions_intel
            WHERE 1=1
        """
        params = []

        if user_name:
            query += " AND user_name = ?"
            params.append(user_name)

        query += " GROUP BY fiscal_year, category, txn_type ORDER BY fiscal_year DESC, category"

        cursor = self.conn.execute(query, params)
        rows = cursor.fetchall()

        # Headers
        headers = ["Fiscal Year", "Category", "Type", "Count", "Total Amount"]
        self._write_headers(ws, headers)

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row["fiscal_year"])
            ws.cell(row=row_idx, column=2, value=row["category"])
            ws.cell(row=row_idx, column=3, value=row["txn_type"])
            ws.cell(row=row_idx, column=4, value=row["txn_count"])

            amount_cell = ws.cell(row=row_idx, column=5, value=row["total_amount"] or 0)
            amount_cell.number_format = '#,##0.00'

            # Color for income/expense
            if row["total_amount"] and row["total_amount"] > 0:
                amount_cell.font = Font(color="006600")
            elif row["total_amount"] and row["total_amount"] < 0:
                amount_cell.font = Font(color="CC0000")

        # Add SUBTOTAL
        if len(rows) > 0:
            subtotal_row = len(rows) + 2
            ws.cell(row=subtotal_row, column=4, value="SUBTOTAL:")
            subtotal_cell = ws.cell(
                row=subtotal_row, column=5,
                value=f"=SUBTOTAL(9,E2:E{len(rows)+1})"
            )
            subtotal_cell.font = Font(bold=True)
            subtotal_cell.number_format = '#,##0.00'

        # Apply auto-filter
        ws.auto_filter.ref = f"A1:E{max(2, len(rows)+1)}"
        ws.freeze_panes = "A2"
        self._adjust_column_widths(ws)

    def _create_category_analysis(
        self,
        wb: "Workbook",
        fiscal_year: Optional[str],
        user_name: Optional[str]
    ) -> None:
        """Create Category Analysis sheet."""
        ws = wb.create_sheet("Category_Analysis")

        # Query by category
        query = """
            SELECT category,
                   COUNT(*) as txn_count,
                   SUM(CASE WHEN txn_type = 'CREDIT' THEN CAST(amount AS REAL) ELSE 0 END) as total_credits,
                   SUM(CASE WHEN txn_type = 'DEBIT' THEN CAST(amount AS REAL) ELSE 0 END) as total_debits,
                   SUM(CAST(amount AS REAL)) as net_amount
            FROM bank_transactions_intel
            WHERE 1=1
        """
        params = []

        if fiscal_year:
            query += " AND fiscal_year = ?"
            params.append(fiscal_year)
        if user_name:
            query += " AND user_name = ?"
            params.append(user_name)

        query += " GROUP BY category ORDER BY net_amount DESC"

        cursor = self.conn.execute(query, params)
        rows = cursor.fetchall()

        # Headers
        headers = ["Category", "Transaction Count", "Total Credits", "Total Debits", "Net Amount"]
        self._write_headers(ws, headers)

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row["category"])
            ws.cell(row=row_idx, column=2, value=row["txn_count"])

            credits_cell = ws.cell(row=row_idx, column=3, value=row["total_credits"] or 0)
            credits_cell.number_format = '#,##0.00'
            credits_cell.font = Font(color="006600")

            debits_cell = ws.cell(row=row_idx, column=4, value=abs(row["total_debits"] or 0))
            debits_cell.number_format = '#,##0.00'
            debits_cell.font = Font(color="CC0000")

            net_cell = ws.cell(row=row_idx, column=5, value=row["net_amount"] or 0)
            net_cell.number_format = '#,##0.00'
            if row["net_amount"] and row["net_amount"] > 0:
                net_cell.font = Font(color="006600")
            elif row["net_amount"] and row["net_amount"] < 0:
                net_cell.font = Font(color="CC0000")

        # Add SUBTOTALs
        if len(rows) > 0:
            subtotal_row = len(rows) + 2
            ws.cell(row=subtotal_row, column=2, value="SUBTOTAL:")
            for col in [3, 4, 5]:
                subtotal_cell = ws.cell(
                    row=subtotal_row, column=col,
                    value=f"=SUBTOTAL(9,{get_column_letter(col)}2:{get_column_letter(col)}{len(rows)+1})"
                )
                subtotal_cell.font = Font(bold=True)
                subtotal_cell.number_format = '#,##0.00'

        ws.auto_filter.ref = f"A1:E{max(2, len(rows)+1)}"
        ws.freeze_panes = "A2"
        self._adjust_column_widths(ws)

    def _create_income_summary(
        self,
        wb: "Workbook",
        fiscal_year: Optional[str],
        user_name: Optional[str]
    ) -> None:
        """Create Income Summary sheet for PFAS asset extraction."""
        ws = wb.create_sheet("Income_Summary")

        # Income categories for PFAS
        income_categories = [
            "RENT_INCOME", "SGB_INTEREST", "DIVIDEND", "SAVINGS_INTEREST",
            "FD_INTEREST", "SALARY", "MF_REDEMPTION"
        ]

        # Query income by category
        placeholders = ",".join("?" * len(income_categories))
        query = f"""
            SELECT category, fiscal_year,
                   SUM(CAST(amount AS REAL)) as total_amount,
                   COUNT(*) as txn_count
            FROM bank_transactions_intel
            WHERE category IN ({placeholders})
              AND txn_type = 'CREDIT'
        """
        params = list(income_categories)

        if fiscal_year:
            query += " AND fiscal_year = ?"
            params.append(fiscal_year)
        if user_name:
            query += " AND user_name = ?"
            params.append(user_name)

        query += " GROUP BY category, fiscal_year ORDER BY fiscal_year DESC, total_amount DESC"

        cursor = self.conn.execute(query, params)
        rows = cursor.fetchall()

        # Headers
        headers = ["Income Category", "Fiscal Year", "Total Amount", "Transaction Count", "PFAS Asset Class"]
        self._write_headers(ws, headers)

        # Asset class mapping
        asset_mapping = {
            "RENT_INCOME": "Rental Income",
            "SGB_INTEREST": "SGB Holdings",
            "DIVIDEND": "Stock Dividends",
            "SAVINGS_INTEREST": "Bank Interest",
            "FD_INTEREST": "Fixed Deposits",
            "SALARY": "Salary Income",
            "MF_REDEMPTION": "Mutual Funds"
        }

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row["category"])
            ws.cell(row=row_idx, column=2, value=row["fiscal_year"])

            amount_cell = ws.cell(row=row_idx, column=3, value=row["total_amount"] or 0)
            amount_cell.number_format = '#,##0.00'
            amount_cell.font = Font(color="006600", bold=True)

            ws.cell(row=row_idx, column=4, value=row["txn_count"])
            ws.cell(row=row_idx, column=5, value=asset_mapping.get(row["category"], ""))

        # Add SUBTOTAL
        if len(rows) > 0:
            subtotal_row = len(rows) + 2
            ws.cell(row=subtotal_row, column=2, value="TOTAL INCOME:")
            subtotal_cell = ws.cell(
                row=subtotal_row, column=3,
                value=f"=SUBTOTAL(9,C2:C{len(rows)+1})"
            )
            subtotal_cell.font = Font(bold=True, color="006600")
            subtotal_cell.number_format = '#,##0.00'

        ws.auto_filter.ref = f"A1:E{max(2, len(rows)+1)}"
        ws.freeze_panes = "A2"
        self._adjust_column_widths(ws)

    def _write_headers(self, ws, headers: List[str]) -> None:
        """Write header row with styling."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def _adjust_column_widths(self, ws) -> None:
        """Adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = max(adjusted_width, 10)

    def get_fiscal_years(self) -> List[str]:
        """Get list of available fiscal years."""
        if not self.conn:
            self.connect()

        cursor = self.conn.execute(
            "SELECT DISTINCT fiscal_year FROM bank_transactions_intel ORDER BY fiscal_year DESC"
        )
        return [row[0] for row in cursor.fetchall()]

    def get_users(self) -> List[str]:
        """Get list of available users."""
        if not self.conn:
            self.connect()

        cursor = self.conn.execute(
            "SELECT DISTINCT user_name FROM bank_transactions_intel ORDER BY user_name"
        )
        return [row[0] for row in cursor.fetchall()]

    def get_income_for_pfas(
        self,
        fiscal_year: str,
        user_name: Optional[str] = None
    ) -> Dict[str, Decimal]:
        """
        Get income totals by category for PFAS update.

        Args:
            fiscal_year: Fiscal year to query
            user_name: Optional user filter

        Returns:
            Dictionary of category -> total amount
        """
        if not self.conn:
            self.connect()

        query = """
            SELECT category, SUM(CAST(amount AS REAL)) as total
            FROM bank_transactions_intel
            WHERE fiscal_year = ?
              AND txn_type = 'CREDIT'
              AND category IN ('RENT_INCOME', 'SGB_INTEREST', 'DIVIDEND', 'SAVINGS_INTEREST')
        """
        params = [fiscal_year]

        if user_name:
            query += " AND user_name = ?"
            params.append(user_name)

        query += " GROUP BY category"

        cursor = self.conn.execute(query, params)
        return {row[0]: Decimal(str(row[1])) for row in cursor.fetchall()}


def main():
    """CLI entry point for testing."""
    import sys

    db_path = "Data/Reports/Bank_Intelligence/money_movement.db"
    output_path = "Data/Reports/Bank_Intelligence/Master_Report.xlsx"

    if len(sys.argv) > 1:
        db_path = sys.argv[1]
    if len(sys.argv) > 2:
        output_path = sys.argv[2]

    print(f"Database: {db_path}")
    print(f"Output: {output_path}")

    with FiscalReportGenerator(db_path) as generator:
        # Show available fiscal years
        fiscal_years = generator.get_fiscal_years()
        print(f"\nAvailable fiscal years: {fiscal_years}")

        # Generate report
        output = generator.generate_master_report(output_path)
        print(f"\nGenerated: {output}")


if __name__ == "__main__":
    main()
