"""Stock Holdings Report Generator.

Generates holdings report with current positions and cost basis.
Calculates average purchase price and unrealized gains.
"""

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from pfas.parsers.stock.models import StockHolding, TradeType


@dataclass
class HoldingsReportData:
    """Holdings report data."""

    as_of_date: date
    holdings: list[StockHolding] = field(default_factory=list)
    total_cost: Decimal = Decimal("0")
    total_value: Optional[Decimal] = None
    total_unrealized_gain: Optional[Decimal] = None

    def calculate_totals(self):
        """Calculate total cost, value, and unrealized gain."""
        self.total_cost = sum(h.total_cost for h in self.holdings)
        if all(h.current_value is not None for h in self.holdings):
            self.total_value = sum(h.current_value for h in self.holdings)
            self.total_unrealized_gain = self.total_value - self.total_cost


class StockHoldingsReport:
    """
    Generate Stock Holdings Report with positions and cost basis.

    Calculates current holdings from all buy/sell trades using FIFO
    (First In, First Out) method for cost basis calculation.
    """

    def __init__(self, db_connection: sqlite3.Connection):
        """
        Initialize report generator.

        Args:
            db_connection: Database connection
        """
        self.conn = db_connection

    def generate(self, user_id: int, as_of_date: Optional[date] = None) -> HoldingsReportData:
        """
        Generate holdings report as of a given date.

        Uses FIFO method for cost basis calculation:
        - Buys are added to position at their cost
        - Sells reduce position starting from oldest lots

        Args:
            user_id: User ID
            as_of_date: Date to calculate holdings as of (default: today)

        Returns:
            HoldingsReportData with positions and cost basis
        """
        if as_of_date is None:
            as_of_date = date.today()

        report = HoldingsReportData(as_of_date=as_of_date)

        # Get all trades up to as_of_date
        trades = self._get_trades(user_id, as_of_date)

        # Group trades by symbol and calculate holdings
        holdings_by_symbol = self._calculate_holdings(trades)

        # Convert to StockHolding objects
        for symbol, data in holdings_by_symbol.items():
            if data['quantity'] > 0:
                holding = StockHolding(
                    symbol=symbol,
                    isin=data.get('isin'),
                    quantity=data['quantity'],
                    average_cost=data['average_cost'],
                    total_cost=data['total_cost'],
                    first_purchase_date=data.get('first_purchase_date'),
                )
                report.holdings.append(holding)

        # Sort by symbol
        report.holdings.sort(key=lambda h: h.symbol)
        report.calculate_totals()

        return report

    def _get_trades(self, user_id: int, as_of_date: date) -> list[dict]:
        """
        Get all trades for a user up to a given date.

        Args:
            user_id: User ID
            as_of_date: Date to get trades up to

        Returns:
            List of trade dictionaries
        """
        cursor = self.conn.execute(
            """
            SELECT
                symbol,
                isin,
                trade_date,
                trade_type,
                quantity,
                price,
                amount,
                trade_category
            FROM stock_trades
            WHERE user_id = ?
                AND trade_date <= ?
                AND (trade_category = 'DELIVERY' OR trade_category IS NULL)
            ORDER BY trade_date, id
            """,
            (user_id, as_of_date.isoformat()),
        )

        trades = []
        for row in cursor.fetchall():
            trades.append({
                'symbol': row['symbol'],
                'isin': row['isin'],
                'trade_date': date.fromisoformat(row['trade_date'])
                if isinstance(row['trade_date'], str) else row['trade_date'],
                'trade_type': row['trade_type'],
                'quantity': int(row['quantity']),
                'price': Decimal(str(row['price'])),
                'amount': Decimal(str(row['amount'])),
            })

        return trades

    def _calculate_holdings(self, trades: list[dict]) -> dict:
        """
        Calculate current holdings using FIFO method.

        Args:
            trades: List of trade dictionaries

        Returns:
            Dictionary of holdings by symbol
        """
        holdings = {}

        for trade in trades:
            symbol = trade['symbol']

            if symbol not in holdings:
                holdings[symbol] = {
                    'isin': trade.get('isin'),
                    'quantity': 0,
                    'total_cost': Decimal("0"),
                    'average_cost': Decimal("0"),
                    'lots': [],  # FIFO lots: [(quantity, price, date), ...]
                    'first_purchase_date': None,
                }

            h = holdings[symbol]

            if trade['trade_type'] == 'BUY':
                # Add to holdings
                h['quantity'] += trade['quantity']
                h['total_cost'] += trade['amount']
                h['lots'].append({
                    'quantity': trade['quantity'],
                    'price': trade['price'],
                    'date': trade['trade_date'],
                })
                if h['first_purchase_date'] is None:
                    h['first_purchase_date'] = trade['trade_date']
                # Recalculate average cost
                if h['quantity'] > 0:
                    h['average_cost'] = h['total_cost'] / h['quantity']

            elif trade['trade_type'] == 'SELL':
                # Remove from holdings using FIFO
                sell_qty = trade['quantity']
                cost_removed = Decimal("0")

                while sell_qty > 0 and h['lots']:
                    lot = h['lots'][0]
                    if lot['quantity'] <= sell_qty:
                        # Use entire lot
                        sell_qty -= lot['quantity']
                        cost_removed += lot['quantity'] * lot['price']
                        h['lots'].pop(0)
                    else:
                        # Partial lot
                        lot['quantity'] -= sell_qty
                        cost_removed += sell_qty * lot['price']
                        sell_qty = 0

                h['quantity'] -= trade['quantity']
                h['total_cost'] -= cost_removed

                # Recalculate average cost
                if h['quantity'] > 0:
                    h['average_cost'] = h['total_cost'] / h['quantity']
                else:
                    h['average_cost'] = Decimal("0")
                    h['first_purchase_date'] = None
                    if h['lots']:
                        h['first_purchase_date'] = h['lots'][0]['date']

        return holdings

    def generate_from_trades(
        self,
        trades: list,
        as_of_date: Optional[date] = None
    ) -> HoldingsReportData:
        """
        Generate holdings report from a list of StockTrade objects.

        Useful when you already have parsed trades and don't want to query DB.

        Args:
            trades: List of StockTrade objects
            as_of_date: Date to calculate holdings as of (default: today)

        Returns:
            HoldingsReportData with positions and cost basis
        """
        if as_of_date is None:
            as_of_date = date.today()

        report = HoldingsReportData(as_of_date=as_of_date)

        # Convert StockTrade to dict format
        trade_dicts = []
        for trade in trades:
            if trade.trade_category and trade.trade_category.value != 'DELIVERY':
                continue  # Skip intraday and F&O trades
            if trade.trade_date > as_of_date:
                continue

            trade_dicts.append({
                'symbol': trade.symbol,
                'isin': trade.isin,
                'trade_date': trade.trade_date,
                'trade_type': trade.trade_type.value,
                'quantity': trade.quantity,
                'price': trade.price,
                'amount': trade.amount,
            })

        # Sort by date
        trade_dicts.sort(key=lambda t: t['trade_date'])

        # Calculate holdings
        holdings_by_symbol = self._calculate_holdings(trade_dicts)

        # Convert to StockHolding objects
        for symbol, data in holdings_by_symbol.items():
            if data['quantity'] > 0:
                holding = StockHolding(
                    symbol=symbol,
                    isin=data.get('isin'),
                    quantity=data['quantity'],
                    average_cost=data['average_cost'],
                    total_cost=data['total_cost'],
                    first_purchase_date=data.get('first_purchase_date'),
                )
                report.holdings.append(holding)

        # Sort by symbol
        report.holdings.sort(key=lambda h: h.symbol)
        report.calculate_totals()

        return report

    def export_excel(self, report: HoldingsReportData, output_path: Path) -> Path:
        """
        Export holdings report to Excel.

        Args:
            report: HoldingsReportData from generate()
            output_path: Output file path (.xlsx)

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Holdings"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        money_format = '#,##0.00'

        row = 1

        # Title
        ws.cell(row=row, column=1, value=f"Stock Holdings Report - As of {report.as_of_date.strftime('%d-%b-%Y')}")
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

        # Summary
        ws.cell(row=row, column=1, value="Portfolio Summary")
        ws.cell(row=row, column=1).font = subheader_font
        row += 1

        ws.cell(row=row, column=1, value="Total Holdings:")
        ws.cell(row=row, column=2, value=len(report.holdings))
        row += 1

        ws.cell(row=row, column=1, value="Total Cost Basis:")
        cell = ws.cell(row=row, column=2, value=report.total_cost)
        cell.number_format = money_format
        row += 2

        # Holdings table
        if report.holdings:
            headers = ["Symbol", "ISIN", "Quantity", "Avg Cost", "Total Cost", "First Purchase"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for holding in report.holdings:
                data = [
                    holding.symbol,
                    holding.isin or "",
                    holding.quantity,
                    holding.average_cost,
                    holding.total_cost,
                    holding.first_purchase_date.strftime("%d-%b-%Y") if holding.first_purchase_date else "",
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col in [4, 5]:  # Money columns
                        cell.number_format = money_format
                row += 1

            # Total row
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3, value=sum(h.quantity for h in report.holdings))
            ws.cell(row=row, column=3).font = Font(bold=True)
            cell = ws.cell(row=row, column=5, value=report.total_cost)
            cell.font = Font(bold=True)
            cell.number_format = money_format
        else:
            ws.cell(row=row, column=1, value="No holdings found.")

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 15

        # Save workbook
        output_path = Path(output_path)
        wb.save(output_path)
        return output_path
