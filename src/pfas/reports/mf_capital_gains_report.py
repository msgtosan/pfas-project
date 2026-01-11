"""Mutual Fund Capital Gains Report Generator.

Generates capital gains statements in Excel and PDF formats with quarterly breakdown.
Supports FY 2024-25 tax rates (Budget 2024):
- Equity STCG: 20%
- Equity LTCG: 12.5% with Rs 1.25L exemption
- Debt: Taxed at slab rate
"""

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pfas.parsers.mf.models import AssetClass


@dataclass
class TransactionDetail:
    """Details of a single MF redemption transaction."""

    folio_number: str
    scheme_name: str
    asset_class: str
    redemption_date: date
    units: Decimal
    nav: Decimal
    amount: Decimal
    purchase_date: Optional[date]
    purchase_amount: Decimal
    holding_days: int
    is_long_term: bool
    stcg: Decimal
    ltcg: Decimal


@dataclass
class QuarterlySummary:
    """
    Summary of capital gains for a quarter.

    Quarters in Indian Financial Year:
    - Q1: April-June
    - Q2: July-September
    - Q3: October-December
    - Q4: January-March
    """

    quarter: str  # "Q1", "Q2", "Q3", "Q4"
    start_date: date
    end_date: date

    # Equity
    equity_stcg: Decimal = Decimal("0")
    equity_ltcg: Decimal = Decimal("0")
    equity_transactions: list[TransactionDetail] = field(default_factory=list)

    # Debt
    debt_stcg: Decimal = Decimal("0")
    debt_ltcg: Decimal = Decimal("0")
    debt_transactions: list[TransactionDetail] = field(default_factory=list)

    @property
    def total_stcg(self) -> Decimal:
        """Total STCG for this quarter."""
        return self.equity_stcg + self.debt_stcg

    @property
    def total_ltcg(self) -> Decimal:
        """Total LTCG for this quarter."""
        return self.equity_ltcg + self.debt_ltcg


@dataclass
class FYSummary:
    """Full year capital gains summary."""

    financial_year: str
    quarters: list[QuarterlySummary]

    # Equity totals
    equity_stcg_total: Decimal = Decimal("0")
    equity_ltcg_total: Decimal = Decimal("0")
    equity_ltcg_exemption: Decimal = Decimal("0")
    equity_taxable_ltcg: Decimal = Decimal("0")

    # Debt totals
    debt_stcg_total: Decimal = Decimal("0")
    debt_ltcg_total: Decimal = Decimal("0")

    # Tax rates
    equity_stcg_rate: Decimal = Decimal("20")
    equity_ltcg_rate: Decimal = Decimal("12.5")

    @property
    def total_stcg(self) -> Decimal:
        """Total STCG for the year."""
        return self.equity_stcg_total + self.debt_stcg_total

    @property
    def total_ltcg(self) -> Decimal:
        """Total LTCG for the year."""
        return self.equity_ltcg_total + self.debt_ltcg_total


class MFCapitalGainsReport:
    """
    Generate MF Capital Gains Statement with quarterly breakdown.

    Tax Rules (Budget 2024, effective FY 2024-25):
    - Equity STCG: 20% flat rate
    - Equity LTCG: 12.5% with Rs 1.25 lakh exemption
    - Debt: Taxed at individual slab rate (no special rate)

    Holding Period:
    - Equity: >12 months = Long Term
    - Debt: Always taxed at slab rate (post Apr 2023)
    """

    LTCG_EXEMPTION = Decimal("125000")  # Rs 1.25 lakh

    # Quarter date ranges (for any FY)
    QUARTERS = [
        ("Q1", 4, 1, 6, 30),   # Apr 1 to Jun 30
        ("Q2", 7, 1, 9, 30),   # Jul 1 to Sep 30
        ("Q3", 10, 1, 12, 31), # Oct 1 to Dec 31
        ("Q4", 1, 1, 3, 31),   # Jan 1 to Mar 31
    ]

    def __init__(self, db_connection: sqlite3.Connection):
        """
        Initialize report generator.

        Args:
            db_connection: Database connection
        """
        self.conn = db_connection

    def generate(self, user_id: int, financial_year: str) -> FYSummary:
        """
        Generate capital gains summary with quarterly breakdown.

        Args:
            user_id: User ID
            financial_year: FY in format '2024-25'

        Returns:
            FYSummary with quarterly breakdown
        """
        # Parse FY
        start_year = int(financial_year.split("-")[0])

        # Build quarter date ranges
        quarters = []
        for q_name, start_month, start_day, end_month, end_day in self.QUARTERS:
            if start_month >= 4:  # Apr-Dec is in start_year
                q_start = date(start_year, start_month, start_day)
                q_end = date(start_year, end_month, end_day)
            else:  # Jan-Mar is in start_year + 1
                q_start = date(start_year + 1, start_month, start_day)
                q_end = date(start_year + 1, end_month, end_day)

            quarter = self._get_quarter_data(user_id, q_name, q_start, q_end)
            quarters.append(quarter)

        # Calculate yearly totals
        equity_stcg = sum(q.equity_stcg for q in quarters)
        equity_ltcg = sum(q.equity_ltcg for q in quarters)
        debt_stcg = sum(q.debt_stcg for q in quarters)
        debt_ltcg = sum(q.debt_ltcg for q in quarters)

        # Apply LTCG exemption (only for equity)
        ltcg_exemption = min(equity_ltcg, self.LTCG_EXEMPTION) if equity_ltcg > 0 else Decimal("0")
        taxable_ltcg = max(Decimal("0"), equity_ltcg - ltcg_exemption)

        return FYSummary(
            financial_year=financial_year,
            quarters=quarters,
            equity_stcg_total=equity_stcg,
            equity_ltcg_total=equity_ltcg,
            equity_ltcg_exemption=ltcg_exemption,
            equity_taxable_ltcg=taxable_ltcg,
            debt_stcg_total=debt_stcg,
            debt_ltcg_total=debt_ltcg,
        )

    def _get_quarter_data(
        self, user_id: int, quarter: str, start_date: date, end_date: date
    ) -> QuarterlySummary:
        """
        Get capital gains data for a specific quarter.

        Args:
            user_id: User ID
            quarter: Quarter name (Q1, Q2, Q3, Q4)
            start_date: Quarter start date
            end_date: Quarter end date

        Returns:
            QuarterlySummary for the quarter
        """
        summary = QuarterlySummary(
            quarter=quarter,
            start_date=start_date,
            end_date=end_date,
        )

        # Query redemption transactions for this quarter
        cursor = self.conn.execute(
            """
            SELECT
                mf.folio_number,
                ms.name as scheme_name,
                ms.asset_class,
                mt.date as redemption_date,
                mt.units,
                mt.nav,
                mt.amount,
                mt.purchase_date,
                mt.purchase_amount,
                mt.holding_period_days,
                mt.is_long_term,
                mt.short_term_gain,
                mt.long_term_gain
            FROM mf_transactions mt
            JOIN mf_folios mf ON mt.folio_id = mf.id
            JOIN mf_schemes ms ON mf.scheme_id = ms.id
            WHERE mt.user_id = ?
                AND mt.transaction_type = 'REDEMPTION'
                AND mt.date >= ?
                AND mt.date <= ?
            ORDER BY mt.date, ms.name
            """,
            (user_id, start_date.isoformat(), end_date.isoformat()),
        )

        for row in cursor.fetchall():
            txn = TransactionDetail(
                folio_number=row["folio_number"],
                scheme_name=row["scheme_name"],
                asset_class=row["asset_class"],
                redemption_date=date.fromisoformat(row["redemption_date"])
                if isinstance(row["redemption_date"], str)
                else row["redemption_date"],
                units=Decimal(str(row["units"])),
                nav=Decimal(str(row["nav"])),
                amount=Decimal(str(row["amount"])),
                purchase_date=date.fromisoformat(row["purchase_date"])
                if row["purchase_date"] and isinstance(row["purchase_date"], str)
                else row["purchase_date"],
                purchase_amount=Decimal(str(row["purchase_amount"] or 0)),
                holding_days=row["holding_period_days"] or 0,
                is_long_term=bool(row["is_long_term"]),
                stcg=Decimal(str(row["short_term_gain"] or 0)),
                ltcg=Decimal(str(row["long_term_gain"] or 0)),
            )

            # Add to appropriate category
            if row["asset_class"] == "EQUITY":
                summary.equity_transactions.append(txn)
                summary.equity_stcg += txn.stcg
                summary.equity_ltcg += txn.ltcg
            else:
                summary.debt_transactions.append(txn)
                summary.debt_stcg += txn.stcg
                summary.debt_ltcg += txn.ltcg

        return summary

    def export_excel(self, summary: FYSummary, output_path: Path) -> Path:
        """
        Export capital gains statement to Excel.

        Args:
            summary: FYSummary from generate()
            output_path: Output file path (.xlsx)

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()

        # Create Summary sheet
        self._create_summary_sheet(wb, summary)

        # Create quarterly detail sheets
        for quarter in summary.quarters:
            self._create_quarter_sheet(wb, quarter)

        # Remove default sheet if empty
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        # Save workbook
        output_path = Path(output_path)
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: Workbook, summary: FYSummary) -> None:
        """Create the summary sheet with yearly and quarterly totals."""
        ws = wb.active
        ws.title = "Summary"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        money_format = '#,##0.00'

        row = 1

        # Title
        ws.cell(row=row, column=1, value=f"Mutual Fund Capital Gains Statement - FY {summary.financial_year}")
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

        # Yearly Summary Section
        ws.cell(row=row, column=1, value="YEARLY SUMMARY")
        ws.cell(row=row, column=1).font = subheader_font
        row += 1

        # Equity Summary
        summary_headers = ["Category", "STCG", "LTCG", "Exemption", "Taxable LTCG", "Tax Rate (%)"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Equity row
        equity_data = [
            "Equity",
            summary.equity_stcg_total,
            summary.equity_ltcg_total,
            summary.equity_ltcg_exemption,
            summary.equity_taxable_ltcg,
            summary.equity_stcg_rate,
        ]
        for col, value in enumerate(equity_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 1:
                cell.number_format = money_format
        row += 1

        # Debt row
        debt_data = [
            "Debt",
            summary.debt_stcg_total,
            summary.debt_ltcg_total,
            Decimal("0"),
            summary.debt_ltcg_total,
            "Slab Rate",
        ]
        for col, value in enumerate(debt_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col > 1 and col < 6:
                cell.number_format = money_format
        row += 2

        # Quarterly Breakdown Section
        ws.cell(row=row, column=1, value="QUARTERLY BREAKDOWN")
        ws.cell(row=row, column=1).font = subheader_font
        row += 1

        # Quarterly headers
        q_headers = ["Quarter", "Period", "Equity STCG", "Equity LTCG", "Debt STCG", "Debt LTCG", "Total"]
        for col, header in enumerate(q_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Quarterly data
        for quarter in summary.quarters:
            period = f"{quarter.start_date.strftime('%d-%b-%Y')} to {quarter.end_date.strftime('%d-%b-%Y')}"
            total = quarter.equity_stcg + quarter.equity_ltcg + quarter.debt_stcg + quarter.debt_ltcg
            q_data = [
                quarter.quarter,
                period,
                quarter.equity_stcg,
                quarter.equity_ltcg,
                quarter.debt_stcg,
                quarter.debt_ltcg,
                total,
            ]
            for col, value in enumerate(q_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col > 2:
                    cell.number_format = money_format
            row += 1

        # Total row
        total_data = [
            "TOTAL",
            "",
            summary.equity_stcg_total,
            summary.equity_ltcg_total,
            summary.debt_stcg_total,
            summary.debt_ltcg_total,
            summary.total_stcg + summary.total_ltcg,
        ]
        for col, value in enumerate(total_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.font = Font(bold=True)
            if col > 2:
                cell.number_format = money_format
        row += 2

        # Tax Notes Section
        ws.cell(row=row, column=1, value="TAX NOTES")
        ws.cell(row=row, column=1).font = subheader_font
        row += 1

        notes = [
            f"1. Equity STCG Tax Rate: {summary.equity_stcg_rate}%",
            f"2. Equity LTCG Tax Rate: {summary.equity_ltcg_rate}% (after Rs 1.25 Lakh exemption)",
            "3. Debt Funds: Taxed at individual slab rate (no special rate from FY 2023-24)",
            "4. Long Term: Equity >12 months, Debt >24 months",
        ]
        for note in notes:
            ws.cell(row=row, column=1, value=note)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        for col in ["C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 15

    def _create_quarter_sheet(self, wb: Workbook, quarter: QuarterlySummary) -> None:
        """Create a detail sheet for a specific quarter."""
        ws = wb.create_sheet(title=quarter.quarter)

        # Styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        money_format = '#,##0.00'

        row = 1

        # Title
        period = f"{quarter.start_date.strftime('%d-%b-%Y')} to {quarter.end_date.strftime('%d-%b-%Y')}"
        ws.cell(row=row, column=1, value=f"{quarter.quarter} Capital Gains ({period})")
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 2

        all_transactions = quarter.equity_transactions + quarter.debt_transactions

        if not all_transactions:
            ws.cell(row=row, column=1, value="No redemption transactions in this quarter.")
            return

        # Transaction headers
        txn_headers = [
            "Folio",
            "Scheme",
            "Asset Class",
            "Redemption Date",
            "Units",
            "NAV",
            "Sale Amount",
            "Purchase Date",
            "Purchase Amount",
            "Holding Days",
            "Type",
            "STCG",
            "LTCG",
        ]
        for col, header in enumerate(txn_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Transaction data
        for txn in all_transactions:
            txn_data = [
                txn.folio_number,
                txn.scheme_name,
                txn.asset_class,
                txn.redemption_date.strftime("%d-%b-%Y") if txn.redemption_date else "",
                float(txn.units),
                float(txn.nav),
                txn.amount,
                txn.purchase_date.strftime("%d-%b-%Y") if txn.purchase_date else "",
                txn.purchase_amount,
                txn.holding_days,
                "LTCG" if txn.is_long_term else "STCG",
                txn.stcg,
                txn.ltcg,
            ]
            for col, value in enumerate(txn_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [7, 9, 12, 13]:  # Money columns
                    cell.number_format = money_format
            row += 1

        # Summary row
        row += 1
        ws.cell(row=row, column=1, value="Quarter Summary:")
        ws.cell(row=row, column=1).font = subheader_font
        row += 1

        summary_data = [
            ("Equity STCG:", quarter.equity_stcg),
            ("Equity LTCG:", quarter.equity_ltcg),
            ("Debt STCG:", quarter.debt_stcg),
            ("Debt LTCG:", quarter.debt_ltcg),
            ("Total:", quarter.total_stcg + quarter.total_ltcg),
        ]
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = money_format
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        for col_num in range(4, 14):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

    def export_pdf(self, summary: FYSummary, output_path: Path) -> Optional[Path]:
        """
        Export capital gains statement to PDF.

        Args:
            summary: FYSummary from generate()
            output_path: Output file path (.pdf)

        Returns:
            Path to generated PDF file, or None if PDF generation is not available

        Note:
            Requires reportlab package. Falls back to None if not installed.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
        except ImportError:
            return None

        output_path = Path(output_path)
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=20,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=10,
            spaceBefore=15,
        )

        elements = []

        # Title
        elements.append(
            Paragraph(
                f"Mutual Fund Capital Gains Statement - FY {summary.financial_year}",
                title_style,
            )
        )

        # Yearly Summary Table
        elements.append(Paragraph("Yearly Summary", section_style))

        summary_data = [
            ["Category", "STCG", "LTCG", "Exemption", "Taxable LTCG", "Tax Rate"],
            [
                "Equity",
                f"Rs {summary.equity_stcg_total:,.2f}",
                f"Rs {summary.equity_ltcg_total:,.2f}",
                f"Rs {summary.equity_ltcg_exemption:,.2f}",
                f"Rs {summary.equity_taxable_ltcg:,.2f}",
                f"{summary.equity_stcg_rate}% / {summary.equity_ltcg_rate}%",
            ],
            [
                "Debt",
                f"Rs {summary.debt_stcg_total:,.2f}",
                f"Rs {summary.debt_ltcg_total:,.2f}",
                "N/A",
                f"Rs {summary.debt_ltcg_total:,.2f}",
                "Slab Rate",
            ],
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Quarterly Breakdown Table
        elements.append(Paragraph("Quarterly Breakdown", section_style))

        quarterly_data = [
            ["Quarter", "Period", "Equity STCG", "Equity LTCG", "Debt STCG", "Debt LTCG", "Total"],
        ]
        for q in summary.quarters:
            period = f"{q.start_date.strftime('%d-%b')} to {q.end_date.strftime('%d-%b')}"
            total = q.equity_stcg + q.equity_ltcg + q.debt_stcg + q.debt_ltcg
            quarterly_data.append(
                [
                    q.quarter,
                    period,
                    f"Rs {q.equity_stcg:,.2f}",
                    f"Rs {q.equity_ltcg:,.2f}",
                    f"Rs {q.debt_stcg:,.2f}",
                    f"Rs {q.debt_ltcg:,.2f}",
                    f"Rs {total:,.2f}",
                ]
            )

        # Add total row
        quarterly_data.append(
            [
                "TOTAL",
                "",
                f"Rs {summary.equity_stcg_total:,.2f}",
                f"Rs {summary.equity_ltcg_total:,.2f}",
                f"Rs {summary.debt_stcg_total:,.2f}",
                f"Rs {summary.debt_ltcg_total:,.2f}",
                f"Rs {summary.total_stcg + summary.total_ltcg:,.2f}",
            ]
        )

        quarterly_table = Table(
            quarterly_data,
            colWidths=[0.8 * inch, 1.8 * inch, 1.3 * inch, 1.3 * inch, 1.3 * inch, 1.3 * inch, 1.3 * inch],
        )
        quarterly_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9E2F3")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ]
            )
        )
        elements.append(quarterly_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Notes
        elements.append(Paragraph("Tax Notes", section_style))
        notes = [
            f"1. Equity STCG Tax Rate: {summary.equity_stcg_rate}%",
            f"2. Equity LTCG Tax Rate: {summary.equity_ltcg_rate}% (after Rs 1.25 Lakh exemption)",
            "3. Debt Funds: Taxed at individual slab rate (no special rate from FY 2023-24)",
            "4. Long Term: Equity >12 months, Debt >24 months",
        ]
        for note in notes:
            elements.append(Paragraph(note, styles["Normal"]))

        # Build PDF
        doc.build(elements)
        return output_path
