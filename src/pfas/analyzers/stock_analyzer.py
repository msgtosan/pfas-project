"""
Stock Statement Analyzer - PFAS

Multi-broker stock statement scanning, normalization, ingestion, and reporting.
Supports ICICI Direct, Zerodha, Groww with extensible config-driven architecture.

Usage:
    from pfas.analyzers.stock_analyzer import StockAnalyzer

    analyzer = StockAnalyzer(conn=db_connection, config_path=Path("config/stock_analyzer_config.json"))
    result = analyzer.analyze(user_name="Sanjay", financial_year="2025-26")

    if result.success:
        report_path = analyzer.generate_reports()
        print(f"Report: {report_path}")
"""

import json
import logging
import re
import hashlib
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


# =============================================================================
# Enums & Constants
# =============================================================================

class BrokerType(Enum):
    """Supported stock brokers."""
    ICICIDIRECT = "icicidirect"
    ZERODHA = "zerodha"
    GROWW = "groww"
    UNKNOWN = "unknown"


class StatementType(Enum):
    """Type of stock statement."""
    HOLDINGS = "holdings"
    TRANSACTIONS = "transactions"
    UNKNOWN = "unknown"


class GainType(Enum):
    """Capital gain classification."""
    STCG = "STCG"  # Short-term: <365 days
    LTCG = "LTCG"  # Long-term: >=365 days


# Indian FY Quarters for Advance Tax
INDIAN_FY_QUARTERS = {
    "Q1": {"start_month": 4, "start_day": 1, "end_month": 6, "end_day": 15, "label": "Q1 (Apr 1 - Jun 15)"},
    "Q2": {"start_month": 6, "start_day": 16, "end_month": 9, "end_day": 15, "label": "Q2 (Jun 16 - Sep 15)"},
    "Q3": {"start_month": 9, "start_day": 16, "end_month": 12, "end_day": 15, "label": "Q3 (Sep 16 - Dec 15)"},
    "Q4": {"start_month": 12, "start_day": 16, "end_month": 3, "end_day": 15, "label": "Q4 (Dec 16 - Mar 15)"},
    "Q5": {"start_month": 3, "start_day": 16, "end_month": 3, "end_day": 31, "label": "Q5 (Mar 16 - Mar 31)"},
}


# =============================================================================
# Data Models
# =============================================================================

@dataclass
class NormalizedHolding:
    """Normalized stock holding record."""
    symbol: str
    isin: str
    company_name: str
    sector: Optional[str]

    quantity_held: int
    quantity_lt: int = 0
    quantity_pledged: int = 0
    quantity_blocked: int = 0

    average_buy_price: Decimal = Decimal("0")
    total_cost_basis: Decimal = Decimal("0")

    current_price: Decimal = Decimal("0")
    market_value: Decimal = Decimal("0")

    unrealized_pnl: Decimal = Decimal("0")
    unrealized_pnl_pct: Decimal = Decimal("0")
    price_change_pct: Decimal = Decimal("0")

    broker: BrokerType = BrokerType.UNKNOWN
    demat_account: str = ""
    as_of_date: Optional[date] = None
    source_file: str = ""

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for DB insertion."""
        return {
            "symbol": self.symbol,
            "isin": self.isin,
            "company_name": self.company_name,
            "sector": self.sector,
            "quantity_held": self.quantity_held,
            "quantity_lt": self.quantity_lt,
            "quantity_pledged": self.quantity_pledged,
            "quantity_blocked": self.quantity_blocked,
            "average_buy_price": float(self.average_buy_price),
            "total_cost_basis": float(self.total_cost_basis),
            "current_price": float(self.current_price),
            "market_value": float(self.market_value),
            "unrealized_pnl": float(self.unrealized_pnl),
            "unrealized_pnl_pct": float(self.unrealized_pnl_pct),
            "price_change_pct": float(self.price_change_pct),
            "demat_account": self.demat_account,
            "as_of_date": self.as_of_date.isoformat() if self.as_of_date else None,
            "source_file": self.source_file,
        }


@dataclass
class NormalizedTransaction:
    """Normalized stock transaction record."""
    symbol: str
    isin: str
    company_name: Optional[str] = None

    quantity: int = 0
    buy_date: Optional[date] = None
    sell_date: Optional[date] = None
    holding_period_days: int = 0

    buy_price: Decimal = Decimal("0")
    sell_price: Decimal = Decimal("0")
    buy_value: Decimal = Decimal("0")
    sell_value: Decimal = Decimal("0")

    buy_expenses: Decimal = Decimal("0")
    sell_expenses: Decimal = Decimal("0")
    brokerage: Decimal = Decimal("0")

    # Grandfathering (for LTCG)
    fmv_31jan2018: Optional[Decimal] = None
    grandfathered_price: Optional[Decimal] = None
    is_grandfathered: bool = False

    # Capital Gains
    profit_loss: Decimal = Decimal("0")
    taxable_profit: Decimal = Decimal("0")
    is_long_term: bool = False
    gain_type: GainType = GainType.STCG

    # Tracking
    stt_paid: Decimal = Decimal("0")
    turnover: Decimal = Decimal("0")
    quarter: Optional[str] = None
    financial_year: Optional[str] = None

    broker: BrokerType = BrokerType.UNKNOWN
    source_file: str = ""

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for DB insertion."""
        return {
            "symbol": self.symbol,
            "isin": self.isin,
            "company_name": self.company_name,
            "quantity": self.quantity,
            "buy_date": self.buy_date.isoformat() if self.buy_date else None,
            "sell_date": self.sell_date.isoformat() if self.sell_date else None,
            "holding_period_days": self.holding_period_days,
            "buy_price": float(self.buy_price),
            "sell_price": float(self.sell_price),
            "buy_value": float(self.buy_value),
            "sell_value": float(self.sell_value),
            "buy_expenses": float(self.buy_expenses),
            "sell_expenses": float(self.sell_expenses),
            "fmv_31jan2018": float(self.fmv_31jan2018) if self.fmv_31jan2018 else None,
            "grandfathered_price": float(self.grandfathered_price) if self.grandfathered_price else None,
            "is_grandfathered": self.is_grandfathered,
            "profit_loss": float(self.profit_loss),
            "taxable_profit": float(self.taxable_profit),
            "is_long_term": self.is_long_term,
            "gain_type": self.gain_type.value,
            "stt_paid": float(self.stt_paid),
            "turnover": float(self.turnover),
            "quarter": self.quarter,
            "financial_year": self.financial_year,
            "source_file": self.source_file,
        }


@dataclass
class ScannedFile:
    """Metadata for a scanned statement file."""
    path: Path
    broker: BrokerType
    statement_type: StatementType
    detection_method: str
    file_hash: str = ""
    modified_date: Optional[datetime] = None

    def __post_init__(self):
        if not self.file_hash and self.path.exists():
            self.file_hash = self._compute_hash()
        if not self.modified_date and self.path.exists():
            self.modified_date = datetime.fromtimestamp(self.path.stat().st_mtime)

    def _compute_hash(self) -> str:
        """Compute MD5 hash of file for deduplication."""
        md5 = hashlib.md5()
        with open(self.path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                md5.update(chunk)
        return md5.hexdigest()


@dataclass
class AnalysisResult:
    """Result of stock analysis."""
    success: bool = True
    files_scanned: int = 0
    holdings_processed: int = 0
    transactions_processed: int = 0
    duplicates_skipped: int = 0

    total_market_value: Decimal = Decimal("0")
    total_cost_basis: Decimal = Decimal("0")
    total_unrealized_pnl: Decimal = Decimal("0")

    total_stcg: Decimal = Decimal("0")
    total_ltcg: Decimal = Decimal("0")
    ltcg_exemption_used: Decimal = Decimal("0")

    by_broker: Dict[str, Dict[str, Any]] = field(default_factory=dict)
    by_quarter: Dict[str, Dict[str, Decimal]] = field(default_factory=dict)

    xirr_overall: Optional[float] = None
    xirr_by_stock: Dict[str, float] = field(default_factory=dict)

    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# =============================================================================
# Broker Detection
# =============================================================================

class BrokerDetector:
    """
    Detects broker from folder structure, filename, and file content.

    Detection priority:
    1. Folder name hints (ICICIDirect/, Zerodha/)
    2. Filename keywords
    3. Content analysis (column signatures, keywords in file)
    """

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.broker_config = config.get("brokers", {})
        self.detection_config = config.get("broker_detection", {})

    def detect(self, file_path: Path) -> Tuple[BrokerType, str]:
        """
        Detect broker for a file.

        Returns:
            Tuple of (BrokerType, detection_method)
        """
        # 1. Check folder structure
        broker, method = self._detect_from_folder(file_path)
        if broker != BrokerType.UNKNOWN:
            return broker, method

        # 2. Check filename
        broker, method = self._detect_from_filename(file_path)
        if broker != BrokerType.UNKNOWN:
            return broker, method

        # 3. Check file content
        broker, method = self._detect_from_content(file_path)
        if broker != BrokerType.UNKNOWN:
            return broker, method

        return BrokerType.UNKNOWN, "undetected"

    def _detect_from_folder(self, file_path: Path) -> Tuple[BrokerType, str]:
        """Detect broker from folder path (directory names only, not filename)."""
        # Only check parent directories, not the filename
        dir_parts = [p.lower() for p in file_path.parent.parts]

        for broker_key, broker_cfg in self.broker_config.items():
            folder_patterns = broker_cfg.get("folder_patterns", [])
            for pattern in folder_patterns:
                pattern_lower = pattern.lower()
                if any(pattern_lower in part for part in dir_parts):
                    return BrokerType(broker_key), f"folder:{pattern}"

        return BrokerType.UNKNOWN, ""

    def _detect_from_filename(self, file_path: Path) -> Tuple[BrokerType, str]:
        """Detect broker from filename keywords."""
        filename = file_path.name.lower()
        # Normalize: replace underscores/hyphens with spaces for flexible matching
        filename_normalized = filename.replace("_", " ").replace("-", " ")

        keywords = self.detection_config.get("keywords", {})
        for broker_key, kw_list in keywords.items():
            for kw in kw_list:
                kw_lower = kw.lower()
                kw_normalized = kw_lower.replace("_", " ").replace("-", " ")
                # Check both original and normalized versions
                if kw_lower in filename or kw_normalized in filename_normalized:
                    return BrokerType(broker_key), f"filename:{kw}"

        return BrokerType.UNKNOWN, ""

    def _detect_from_content(self, file_path: Path) -> Tuple[BrokerType, str]:
        """Detect broker from file content (column names, keywords)."""
        suffix = file_path.suffix.lower()
        if suffix not in [".xlsx", ".xls", ".csv"]:
            return BrokerType.UNKNOWN, ""

        try:
            if suffix == ".csv":
                df = pd.read_csv(file_path, nrows=10)
            else:
                df = pd.read_excel(file_path, nrows=10)

            columns = " ".join(df.columns.astype(str)).lower()

            # Check column signatures
            col_signatures = self.detection_config.get("column_signatures", {})
            for broker_key, signatures in col_signatures.items():
                for sig in signatures:
                    if sig.lower() in columns:
                        return BrokerType(broker_key), f"content:column:{sig}"

            # Check content keywords
            content_str = df.to_string().lower()
            keywords = self.detection_config.get("keywords", {})
            for broker_key, kw_list in keywords.items():
                for kw in kw_list:
                    if kw.lower() in content_str:
                        return BrokerType(broker_key), f"content:keyword:{kw}"

        except Exception as e:
            logger.debug(f"Content detection failed for {file_path}: {e}")

        return BrokerType.UNKNOWN, ""

    def detect_statement_type(self, file_path: Path, broker: BrokerType) -> StatementType:
        """Detect if file is holdings or transactions statement."""
        filename = file_path.name.lower()

        broker_key = broker.value if broker != BrokerType.UNKNOWN else None

        # Check file patterns from config
        for br_key, br_cfg in self.broker_config.items():
            if broker_key and br_key != broker_key:
                continue

            holdings_patterns = br_cfg.get("file_patterns", {}).get("holdings", [])
            for pattern in holdings_patterns:
                # Convert glob pattern to regex-ish check
                pattern_re = pattern.replace("*", ".*").lower()
                if re.search(pattern_re, filename):
                    return StatementType.HOLDINGS

            txn_patterns = br_cfg.get("file_patterns", {}).get("transactions", [])
            for pattern in txn_patterns:
                pattern_re = pattern.replace("*", ".*").lower()
                if re.search(pattern_re, filename):
                    return StatementType.TRANSACTIONS

        # Fallback: keyword-based detection
        holdings_keywords = ["holding", "portfolio", "demat", "position"]
        txn_keywords = ["capital_gain", "cg_", "taxpnl", "tax_pnl", "trade", "pnl"]

        for kw in holdings_keywords:
            if kw in filename:
                return StatementType.HOLDINGS

        for kw in txn_keywords:
            if kw in filename:
                return StatementType.TRANSACTIONS

        return StatementType.UNKNOWN


# =============================================================================
# Statement Scanner
# =============================================================================

class StockStatementScanner:
    """
    Recursively scans folders for stock statement files.

    Supports ICICI Direct, Zerodha, Groww with automatic broker detection.
    """

    def __init__(self, config: Dict[str, Any], path_resolver=None):
        self.config = config
        self.path_resolver = path_resolver
        self.broker_detector = BrokerDetector(config)
        self.seen_hashes: set = set()

    def scan(
        self,
        base_path: Path,
        recursive: bool = True,
        include_archive: bool = False
    ) -> List[ScannedFile]:
        """
        Scan folder for stock statement files.

        Args:
            base_path: Root folder to scan (e.g., inbox/Indian-Stocks)
            recursive: Whether to scan subfolders
            include_archive: Whether to include archive folder

        Returns:
            List of ScannedFile objects
        """
        scanned_files: List[ScannedFile] = []
        self.seen_hashes.clear()

        paths_to_scan = [base_path]
        if include_archive and self.path_resolver:
            archive_path = self.path_resolver.archive() / "Indian-Stocks"
            if archive_path.exists():
                paths_to_scan.append(archive_path)

        extensions = {".xlsx", ".xls", ".csv"}
        exclude_patterns = ["~$", ".tmp", "failed"]

        for scan_path in paths_to_scan:
            if not scan_path.exists():
                logger.warning(f"Path does not exist: {scan_path}")
                continue

            pattern = "**/*" if recursive else "*"
            for file_path in scan_path.glob(pattern):
                if not file_path.is_file():
                    continue

                # Check extension
                if file_path.suffix.lower() not in extensions:
                    continue

                # Check exclude patterns
                if any(excl in file_path.name for excl in exclude_patterns):
                    continue

                # Detect broker and statement type
                broker, detection_method = self.broker_detector.detect(file_path)
                stmt_type = self.broker_detector.detect_statement_type(file_path, broker)

                scanned = ScannedFile(
                    path=file_path,
                    broker=broker,
                    statement_type=stmt_type,
                    detection_method=detection_method
                )

                # Skip duplicates
                if scanned.file_hash in self.seen_hashes:
                    logger.debug(f"Skipping duplicate: {file_path.name}")
                    continue

                self.seen_hashes.add(scanned.file_hash)
                scanned_files.append(scanned)

        logger.info(f"Scanned {len(scanned_files)} files from {base_path}")
        return scanned_files


# =============================================================================
# Field Normalizer
# =============================================================================

class StockFieldNormalizer:
    """
    Normalizes broker-specific fields to canonical schema.

    Handles ICICI Direct, Zerodha, Groww field mappings and value transformations.
    """

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.field_mappings = config.get("field_mappings", {})
        self.value_transforms = config.get("value_transforms", {})

    def normalize_holdings(
        self,
        df: pd.DataFrame,
        broker: BrokerType,
        source_file: str
    ) -> List[NormalizedHolding]:
        """
        Normalize holdings DataFrame to canonical format.

        Args:
            df: Raw DataFrame from broker statement
            broker: Detected broker type
            source_file: Source file path for tracking

        Returns:
            List of NormalizedHolding objects
        """
        broker_key = broker.value
        mappings = self.field_mappings.get("holdings", {}).get(broker_key, {})

        if not mappings:
            logger.warning(f"No holdings mapping for broker: {broker_key}")
            return []

        holdings = []
        for _, row in df.iterrows():
            try:
                holding = self._normalize_holding_row(row, mappings, broker, source_file)
                if holding and holding.quantity_held > 0:
                    holdings.append(holding)
            except Exception as e:
                logger.warning(f"Failed to normalize holding row: {e}")

        return holdings

    def _normalize_holding_row(
        self,
        row: pd.Series,
        mappings: Dict[str, str],
        broker: BrokerType,
        source_file: str
    ) -> Optional[NormalizedHolding]:
        """Normalize a single holdings row."""
        # Build normalized dict
        norm = {}
        for src_col, tgt_field in mappings.items():
            if src_col in row.index:
                norm[tgt_field] = row[src_col]

        # Required fields
        symbol = self._get_str(norm.get("symbol", norm.get("company_name", "")))
        isin = self._get_str(norm.get("isin", ""))
        company_name = self._get_str(norm.get("company_name", symbol))

        if not symbol and not isin:
            return None

        # Quantities
        quantity_held = self._get_int(norm.get("quantity_held", 0))
        quantity_lt = self._get_int(norm.get("quantity_lt", 0))
        quantity_pledged = self._get_int(
            norm.get("quantity_pledged", 0) or
            norm.get("quantity_pledged_margin", 0) or 0
        )
        if "quantity_pledged_loan" in norm:
            quantity_pledged += self._get_int(norm.get("quantity_pledged_loan", 0))

        quantity_blocked = self._get_int(norm.get("quantity_blocked_trade", 0))

        # Prices & Values
        average_buy_price = self._get_decimal(norm.get("average_buy_price", 0))
        current_price = self._get_decimal(norm.get("current_price", 0))
        market_value = self._get_decimal(norm.get("market_value", 0))

        # Calculate missing values
        if market_value == Decimal("0") and quantity_held > 0 and current_price > 0:
            market_value = current_price * quantity_held

        total_cost_basis = average_buy_price * quantity_held

        # P&L
        unrealized_pnl = self._get_decimal(norm.get("unrealized_pnl", 0))
        if unrealized_pnl == Decimal("0") and market_value > 0 and total_cost_basis > 0:
            unrealized_pnl = market_value - total_cost_basis

        unrealized_pnl_pct = self._get_decimal(norm.get("unrealized_pnl_pct", 0))
        if unrealized_pnl_pct == Decimal("0") and total_cost_basis > 0:
            unrealized_pnl_pct = (unrealized_pnl / total_cost_basis) * 100

        price_change_pct = self._get_decimal(norm.get("price_change_pct", 0))

        return NormalizedHolding(
            symbol=symbol,
            isin=isin,
            company_name=company_name,
            sector=self._get_str(norm.get("sector")),
            quantity_held=quantity_held,
            quantity_lt=quantity_lt,
            quantity_pledged=quantity_pledged,
            quantity_blocked=quantity_blocked,
            average_buy_price=average_buy_price,
            total_cost_basis=total_cost_basis,
            current_price=current_price,
            market_value=market_value,
            unrealized_pnl=unrealized_pnl,
            unrealized_pnl_pct=unrealized_pnl_pct,
            price_change_pct=price_change_pct,
            broker=broker,
            demat_account=broker.value.upper(),
            source_file=source_file
        )

    def normalize_transactions(
        self,
        df: pd.DataFrame,
        broker: BrokerType,
        source_file: str,
        financial_year: Optional[str] = None
    ) -> List[NormalizedTransaction]:
        """
        Normalize transactions DataFrame to canonical format.

        Args:
            df: Raw DataFrame from broker statement
            broker: Detected broker type
            source_file: Source file path
            financial_year: Target FY (e.g., "2025-26")

        Returns:
            List of NormalizedTransaction objects
        """
        broker_key = broker.value
        mappings = self.field_mappings.get("transactions", {}).get(broker_key, {})

        if not mappings:
            logger.warning(f"No transaction mapping for broker: {broker_key}")
            return []

        transactions = []
        for _, row in df.iterrows():
            try:
                txn = self._normalize_transaction_row(
                    row, mappings, broker, source_file, financial_year
                )
                if txn and txn.quantity > 0:
                    transactions.append(txn)
            except Exception as e:
                logger.warning(f"Failed to normalize transaction row: {e}")

        return transactions

    def _normalize_transaction_row(
        self,
        row: pd.Series,
        mappings: Dict[str, str],
        broker: BrokerType,
        source_file: str,
        financial_year: Optional[str]
    ) -> Optional[NormalizedTransaction]:
        """Normalize a single transaction row."""
        # Build normalized dict
        norm = {}
        for src_col, tgt_field in mappings.items():
            if src_col in row.index:
                norm[tgt_field] = row[src_col]

        # Required fields
        symbol = self._get_str(norm.get("symbol", ""))
        isin = self._get_str(norm.get("isin", ""))

        if not symbol and not isin:
            return None

        # Dates
        buy_date = self._get_date(norm.get("buy_date"))
        sell_date = self._get_date(norm.get("sell_date"))

        if not sell_date:
            return None  # Need sell date for capital gains

        # Calculate holding period
        holding_period_days = self._get_int(norm.get("holding_period_days", 0))
        if holding_period_days == 0 and buy_date and sell_date:
            holding_period_days = (sell_date - buy_date).days

        # Quantity and prices
        quantity = self._get_int(norm.get("quantity", 0))
        buy_price = self._get_decimal(norm.get("buy_price", 0))
        sell_price = self._get_decimal(norm.get("sell_price", 0))
        buy_value = self._get_decimal(norm.get("buy_value", 0))
        sell_value = self._get_decimal(norm.get("sell_value", 0))

        # Calculate missing values
        if buy_value == Decimal("0") and quantity > 0 and buy_price > 0:
            buy_value = buy_price * quantity
        if sell_value == Decimal("0") and quantity > 0 and sell_price > 0:
            sell_value = sell_price * quantity
        if buy_price == Decimal("0") and quantity > 0 and buy_value > 0:
            buy_price = buy_value / quantity
        if sell_price == Decimal("0") and quantity > 0 and sell_value > 0:
            sell_price = sell_value / quantity

        # Expenses
        buy_expenses = self._get_decimal(norm.get("buy_expenses", 0))
        sell_expenses = self._get_decimal(norm.get("sell_expenses", 0))
        brokerage = self._get_decimal(norm.get("brokerage", 0))

        # Grandfathering
        fmv_31jan2018 = self._get_decimal(norm.get("fmv_31jan2018")) if norm.get("fmv_31jan2018") else None
        grandfathered_price = self._get_decimal(norm.get("grandfathered_price")) if norm.get("grandfathered_price") else None
        is_grandfathered = grandfathered_price is not None and grandfathered_price != buy_price

        # Capital gains
        profit_loss = self._get_decimal(norm.get("profit_loss", 0))
        if profit_loss == Decimal("0"):
            profit_loss = sell_value - buy_value - buy_expenses - sell_expenses

        taxable_profit = self._get_decimal(norm.get("taxable_profit", profit_loss))

        # Long-term classification (365 days for equity)
        is_long_term = holding_period_days >= 365
        gain_type = GainType.LTCG if is_long_term else GainType.STCG

        # STT and turnover
        stt_paid = self._get_decimal(norm.get("stcg_stt_paid", 0))
        turnover = self._get_decimal(norm.get("turnover", 0))
        if turnover == Decimal("0"):
            turnover = sell_value

        # Quarter determination
        quarter = self._determine_quarter(sell_date)

        # FY determination
        if not financial_year and sell_date:
            financial_year = self._determine_fy(sell_date)

        return NormalizedTransaction(
            symbol=symbol,
            isin=isin,
            company_name=self._get_str(norm.get("company_name")),
            quantity=quantity,
            buy_date=buy_date,
            sell_date=sell_date,
            holding_period_days=holding_period_days,
            buy_price=buy_price,
            sell_price=sell_price,
            buy_value=buy_value,
            sell_value=sell_value,
            buy_expenses=buy_expenses,
            sell_expenses=sell_expenses,
            brokerage=brokerage,
            fmv_31jan2018=fmv_31jan2018,
            grandfathered_price=grandfathered_price,
            is_grandfathered=is_grandfathered,
            profit_loss=profit_loss,
            taxable_profit=taxable_profit,
            is_long_term=is_long_term,
            gain_type=gain_type,
            stt_paid=stt_paid,
            turnover=turnover,
            quarter=quarter,
            financial_year=financial_year,
            broker=broker,
            source_file=source_file
        )

    def _determine_quarter(self, sell_date: date) -> Optional[str]:
        """Determine Indian FY quarter from sell date."""
        if not sell_date:
            return None

        month, day = sell_date.month, sell_date.day

        for qtr, bounds in INDIAN_FY_QUARTERS.items():
            start_m, start_d = bounds["start_month"], bounds["start_day"]
            end_m, end_d = bounds["end_month"], bounds["end_day"]

            # Handle Q4 which spans Dec-Mar (two calendar years)
            if start_m > end_m:
                # e.g., Dec 16 to Mar 15
                if (month == start_m and day >= start_d) or \
                   (month > start_m) or \
                   (month < end_m) or \
                   (month == end_m and day <= end_d):
                    return qtr
            else:
                if (month > start_m or (month == start_m and day >= start_d)) and \
                   (month < end_m or (month == end_m and day <= end_d)):
                    return qtr

        return None

    def _determine_fy(self, d: date) -> str:
        """Determine financial year from date."""
        if d.month >= 4:
            return f"{d.year}-{str(d.year + 1)[-2:]}"
        else:
            return f"{d.year - 1}-{str(d.year)[-2:]}"

    # Value conversion helpers
    def _get_str(self, val: Any) -> str:
        """Convert value to string."""
        if pd.isna(val) or val is None:
            return ""
        return str(val).strip()

    def _get_int(self, val: Any) -> int:
        """Convert value to integer."""
        if pd.isna(val) or val is None:
            return 0
        try:
            # Remove commas and parse
            cleaned = str(val).replace(",", "").replace(" ", "").strip()
            return int(float(cleaned))
        except (ValueError, TypeError):
            return 0

    def _get_decimal(self, val: Any) -> Decimal:
        """Convert value to Decimal."""
        if pd.isna(val) or val is None:
            return Decimal("0")
        try:
            # Clean currency symbols and commas
            cleaned = str(val)
            for char in ["₹", "Rs.", "Rs", "INR", ",", " "]:
                cleaned = cleaned.replace(char, "")
            cleaned = cleaned.strip()
            if cleaned == "" or cleaned == "-":
                return Decimal("0")
            return Decimal(cleaned)
        except InvalidOperation:
            return Decimal("0")

    def _get_date(self, val: Any) -> Optional[date]:
        """Convert value to date."""
        if pd.isna(val) or val is None:
            return None

        # If already a date/datetime
        if isinstance(val, date):
            return val
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, pd.Timestamp):
            return val.date()

        # Parse string
        val_str = str(val).strip()
        formats = [
            "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
            "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d-%B-%Y"
        ]

        for fmt in formats:
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue

        return None


# =============================================================================
# Database Ingester
# =============================================================================

class StockDBIngester:
    """
    Idempotent database ingestion for stock holdings and transactions.

    Uses unique constraints to avoid duplicates and upsert logic for updates.
    """

    def __init__(self, conn, config: Dict[str, Any]):
        self.conn = conn
        self.config = config
        self.processing = config.get("processing", {})

    def get_or_create_broker(self, broker: BrokerType) -> int:
        """Get or create broker ID."""
        cursor = self.conn.execute(
            "SELECT id FROM stock_brokers WHERE broker_code = ?",
            (broker.value.upper(),)
        )
        row = cursor.fetchone()
        if row:
            return row[0]

        # Insert new broker
        cursor = self.conn.execute(
            "INSERT INTO stock_brokers (name, broker_code) VALUES (?, ?)",
            (broker.value.title(), broker.value.upper())
        )
        self.conn.commit()
        return cursor.lastrowid

    def ingest_holdings(
        self,
        holdings: List[NormalizedHolding],
        user_id: int,
        as_of_date: date
    ) -> Tuple[int, int]:
        """
        Ingest holdings into database.

        Returns:
            Tuple of (inserted_count, skipped_count)
        """
        inserted = 0
        skipped = 0

        for holding in holdings:
            broker_id = self.get_or_create_broker(holding.broker)

            # Check for existing record
            cursor = self.conn.execute(
                """SELECT id FROM stock_holdings
                   WHERE user_id = ? AND broker_id = ? AND isin = ? AND as_of_date = ?""",
                (user_id, broker_id, holding.isin, as_of_date.isoformat())
            )

            if cursor.fetchone():
                # Update existing
                if self._update_holding(holding, user_id, broker_id, as_of_date):
                    inserted += 1
                else:
                    skipped += 1
            else:
                # Insert new
                self._insert_holding(holding, user_id, broker_id, as_of_date)
                inserted += 1

        self.conn.commit()
        return inserted, skipped

    def _insert_holding(
        self,
        holding: NormalizedHolding,
        user_id: int,
        broker_id: int,
        as_of_date: date
    ):
        """Insert a new holding record."""
        self.conn.execute(
            """INSERT INTO stock_holdings (
                user_id, broker_id, symbol, isin, company_name, sector,
                quantity_held, quantity_lt, quantity_pledged, quantity_blocked,
                average_buy_price, total_cost_basis, current_price, market_value,
                unrealized_pnl, unrealized_pnl_pct, price_change_pct,
                as_of_date, source_file, demat_account
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                user_id, broker_id, holding.symbol, holding.isin, holding.company_name,
                holding.sector, holding.quantity_held, holding.quantity_lt,
                holding.quantity_pledged, holding.quantity_blocked,
                float(holding.average_buy_price), float(holding.total_cost_basis),
                float(holding.current_price), float(holding.market_value),
                float(holding.unrealized_pnl), float(holding.unrealized_pnl_pct),
                float(holding.price_change_pct),
                as_of_date.isoformat(), holding.source_file, holding.demat_account
            )
        )

    def _update_holding(
        self,
        holding: NormalizedHolding,
        user_id: int,
        broker_id: int,
        as_of_date: date
    ) -> bool:
        """Update existing holding record. Returns True if updated."""
        self.conn.execute(
            """UPDATE stock_holdings SET
                symbol = ?, company_name = ?, sector = ?,
                quantity_held = ?, quantity_lt = ?, quantity_pledged = ?, quantity_blocked = ?,
                average_buy_price = ?, total_cost_basis = ?, current_price = ?, market_value = ?,
                unrealized_pnl = ?, unrealized_pnl_pct = ?, price_change_pct = ?,
                source_file = ?, demat_account = ?, updated_at = CURRENT_TIMESTAMP
            WHERE user_id = ? AND broker_id = ? AND isin = ? AND as_of_date = ?""",
            (
                holding.symbol, holding.company_name, holding.sector,
                holding.quantity_held, holding.quantity_lt, holding.quantity_pledged,
                holding.quantity_blocked, float(holding.average_buy_price),
                float(holding.total_cost_basis), float(holding.current_price),
                float(holding.market_value), float(holding.unrealized_pnl),
                float(holding.unrealized_pnl_pct), float(holding.price_change_pct),
                holding.source_file, holding.demat_account,
                user_id, broker_id, holding.isin, as_of_date.isoformat()
            )
        )
        return True

    def ingest_transactions(
        self,
        transactions: List[NormalizedTransaction],
        user_id: int
    ) -> Tuple[int, int]:
        """
        Ingest transactions into database.

        Returns:
            Tuple of (inserted_count, skipped_count)
        """
        inserted = 0
        skipped = 0

        duplicate_key = self.processing.get(
            "duplicate_key",
            ["user_id", "isin", "buy_date", "sell_date", "quantity"]
        )

        for txn in transactions:
            broker_id = self.get_or_create_broker(txn.broker)

            # Check for duplicate
            if self._is_duplicate_transaction(txn, user_id, duplicate_key):
                skipped += 1
                continue

            self._insert_transaction(txn, user_id, broker_id)
            inserted += 1

        self.conn.commit()
        return inserted, skipped

    def _is_duplicate_transaction(
        self,
        txn: NormalizedTransaction,
        user_id: int,
        duplicate_key: List[str]
    ) -> bool:
        """Check if transaction already exists."""
        cursor = self.conn.execute(
            """SELECT id FROM stock_capital_gains_detail
               WHERE user_id = ? AND symbol = ? AND buy_date = ? AND sell_date = ? AND quantity = ?""",
            (
                user_id,
                txn.symbol,
                txn.buy_date.isoformat() if txn.buy_date else None,
                txn.sell_date.isoformat() if txn.sell_date else None,
                txn.quantity
            )
        )
        return cursor.fetchone() is not None

    def _insert_transaction(
        self,
        txn: NormalizedTransaction,
        user_id: int,
        broker_id: int
    ):
        """Insert a new transaction record."""
        self.conn.execute(
            """INSERT INTO stock_capital_gains_detail (
                user_id, broker_id, financial_year, quarter,
                symbol, isin, company_name, quantity,
                buy_date, sell_date, holding_period_days,
                buy_price, sell_price, buy_value, sell_value,
                buy_expenses, sell_expenses,
                fmv_31jan2018, grandfathered_price, is_grandfathered,
                gross_profit_loss, cost_of_acquisition, taxable_profit,
                is_long_term, gain_type, stt_paid, turnover, source_file
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                user_id, broker_id, txn.financial_year, txn.quarter,
                txn.symbol, txn.isin, txn.company_name, txn.quantity,
                txn.buy_date.isoformat() if txn.buy_date else None,
                txn.sell_date.isoformat() if txn.sell_date else None,
                txn.holding_period_days,
                float(txn.buy_price), float(txn.sell_price),
                float(txn.buy_value), float(txn.sell_value),
                float(txn.buy_expenses), float(txn.sell_expenses),
                float(txn.fmv_31jan2018) if txn.fmv_31jan2018 else None,
                float(txn.grandfathered_price) if txn.grandfathered_price else None,
                txn.is_grandfathered,
                float(txn.profit_loss), float(txn.buy_value + txn.buy_expenses),
                float(txn.taxable_profit),
                txn.is_long_term, txn.gain_type.value,
                float(txn.stt_paid), float(txn.turnover), txn.source_file
            )
        )


# =============================================================================
# XIRR Calculator
# =============================================================================

class XIRRCalculator:
    """
    Calculate XIRR (Extended Internal Rate of Return) for stock investments.

    Uses Newton-Raphson method for iterative solving.
    """

    @staticmethod
    def calculate(
        cashflows: List[Tuple[date, Decimal]],
        guess: float = 0.1,
        max_iterations: int = 100,
        tolerance: float = 1e-6
    ) -> Optional[float]:
        """
        Calculate XIRR for a series of cashflows.

        Args:
            cashflows: List of (date, amount) tuples. Negative = outflow, Positive = inflow
            guess: Initial guess for rate
            max_iterations: Max iterations for convergence
            tolerance: Convergence tolerance

        Returns:
            Annual rate as decimal (0.15 = 15%), or None if no convergence
        """
        if len(cashflows) < 2:
            return None

        # Sort by date
        sorted_cf = sorted(cashflows, key=lambda x: x[0])
        dates = [cf[0] for cf in sorted_cf]
        amounts = [float(cf[1]) for cf in sorted_cf]

        # Need at least one positive and one negative cashflow
        if not (any(a < 0 for a in amounts) and any(a > 0 for a in amounts)):
            return None

        base_date = dates[0]
        days = [(d - base_date).days for d in dates]

        rate = guess

        for _ in range(max_iterations):
            # Calculate NPV and derivative
            npv = 0.0
            dnpv = 0.0

            for i, amount in enumerate(amounts):
                t = days[i] / 365.0
                factor = (1 + rate) ** (-t)
                npv += amount * factor
                if t != 0:
                    dnpv -= t * amount * factor / (1 + rate)

            if abs(dnpv) < 1e-10:
                break

            new_rate = rate - npv / dnpv

            if abs(new_rate - rate) < tolerance:
                return new_rate

            rate = new_rate

            # Bound the rate to prevent divergence
            if rate < -0.99:
                rate = -0.99
            if rate > 10:
                rate = 10

        return rate if abs(npv) < 0.01 else None

    @staticmethod
    def calculate_for_stock(
        transactions: List[NormalizedTransaction],
        current_holdings: Optional[NormalizedHolding] = None
    ) -> Optional[float]:
        """
        Calculate XIRR for a specific stock.

        Args:
            transactions: List of transactions for the stock
            current_holdings: Current holding (for unrealized gains)

        Returns:
            XIRR as decimal, or None
        """
        cashflows: List[Tuple[date, Decimal]] = []

        for txn in transactions:
            # Buys are negative (outflow)
            if txn.buy_date and txn.buy_value > 0:
                cashflows.append((txn.buy_date, -txn.buy_value))

            # Sells are positive (inflow)
            if txn.sell_date and txn.sell_value > 0:
                cashflows.append((txn.sell_date, txn.sell_value))

        # Add current holding as final positive cashflow
        if current_holdings and current_holdings.market_value > 0:
            as_of = current_holdings.as_of_date or date.today()
            cashflows.append((as_of, current_holdings.market_value))

        return XIRRCalculator.calculate(cashflows)


# =============================================================================
# Report Generator
# =============================================================================

class StockReportGenerator:
    """
    Generate Excel reports for stock analysis.

    Supports multiple sheets: Portfolio Summary, Holdings Detail, Capital Gains, XIRR.
    """

    def __init__(self, conn, config: Dict[str, Any], result: AnalysisResult):
        self.conn = conn
        self.config = config
        self.result = result
        self.report_settings = config.get("report_settings", {}).get("excel", {})
        self.cg_rules = config.get("capital_gains_rules", {})

    def generate(self, output_path: Path, user_id: int, financial_year: str) -> Path:
        """
        Generate comprehensive Excel report.

        Args:
            output_path: Output file path
            user_id: User ID for filtering
            financial_year: Target FY (e.g., "2025-26")

        Returns:
            Path to generated report
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Generate configured sheets
        sheets_config = self.report_settings.get("sheets", [])

        for sheet_cfg in sheets_config:
            sheet_name = sheet_cfg["name"]
            sheet_type = sheet_cfg["type"]

            if sheet_type == "summary":
                self._add_summary_sheet(wb, user_id)
            elif sheet_type == "holdings":
                self._add_holdings_sheet(wb, user_id)
            elif sheet_type == "sector_allocation":
                self._add_sector_sheet(wb, user_id)
            elif sheet_type == "transactions":
                self._add_transactions_sheet(wb, user_id, financial_year)
            elif sheet_type == "capital_gains_fy":
                self._add_cg_fy_sheet(wb, user_id, financial_year)
            elif sheet_type == "capital_gains_quarterly":
                self._add_cg_quarterly_sheet(wb, user_id, financial_year)
            elif sheet_type == "xirr":
                self._add_xirr_sheet(wb, user_id)
            elif sheet_type == "dividends":
                self._add_dividends_sheet(wb, user_id, financial_year)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Report saved to: {output_path}")

        return output_path

    def _add_summary_sheet(self, wb: Workbook, user_id: int):
        """Add portfolio summary sheet."""
        ws = wb.create_sheet("Portfolio Summary")

        # Fetch aggregated data
        cursor = self.conn.execute(
            """SELECT
                COUNT(DISTINCT symbol) as stock_count,
                SUM(market_value) as total_market_value,
                SUM(total_cost_basis) as total_cost,
                SUM(unrealized_pnl) as total_pnl,
                COUNT(DISTINCT sector) as sector_count
            FROM stock_holdings
            WHERE user_id = ?""",
            (user_id,)
        )
        row = cursor.fetchone()

        # Summary metrics
        metrics = [
            ("Portfolio Summary", ""),
            ("", ""),
            ("Total Stocks Held", row[0] if row else 0),
            ("Sectors", row[4] if row else 0),
            ("", ""),
            ("Total Market Value", f"₹{row[1]:,.2f}" if row and row[1] else "₹0.00"),
            ("Total Cost Basis", f"₹{row[2]:,.2f}" if row and row[2] else "₹0.00"),
            ("Unrealized P&L", f"₹{row[3]:,.2f}" if row and row[3] else "₹0.00"),
            ("Unrealized P&L %", f"{((row[3] / row[2]) * 100):.2f}%" if row and row[2] and row[2] > 0 else "0.00%"),
        ]

        # Add XIRR if available
        if self.result.xirr_overall:
            metrics.append(("Overall XIRR", f"{self.result.xirr_overall * 100:.2f}%"))

        # Write metrics
        for row_idx, (label, value) in enumerate(metrics, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        self._apply_header_style(ws, 1)

    def _add_holdings_sheet(self, wb: Workbook, user_id: int):
        """Add detailed holdings sheet."""
        ws = wb.create_sheet("Holdings Detail")

        # Fetch holdings
        cursor = self.conn.execute(
            """SELECT
                b.broker_code, h.symbol, h.isin, h.company_name, h.sector,
                h.quantity_held, h.quantity_lt, h.quantity_pledged,
                h.average_buy_price, h.current_price, h.market_value,
                h.unrealized_pnl, h.unrealized_pnl_pct, h.as_of_date
            FROM stock_holdings h
            LEFT JOIN stock_brokers b ON h.broker_id = b.id
            WHERE h.user_id = ?
            ORDER BY h.market_value DESC""",
            (user_id,)
        )

        # Headers
        headers = [
            "Broker", "Symbol", "ISIN", "Company", "Sector",
            "Qty Held", "Qty LT", "Qty Pledged",
            "Avg Buy Price", "Current Price", "Market Value",
            "Unrealized P&L", "P&L %", "As Of Date"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1)

        # Data rows
        for row_idx, row_data in enumerate(cursor.fetchall(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _add_sector_sheet(self, wb: Workbook, user_id: int):
        """Add sector allocation sheet."""
        ws = wb.create_sheet("Sector Allocation")

        cursor = self.conn.execute(
            """SELECT
                COALESCE(sector, 'Unknown') as sector,
                COUNT(DISTINCT symbol) as stock_count,
                SUM(market_value) as total_value,
                SUM(unrealized_pnl) as total_pnl
            FROM stock_holdings
            WHERE user_id = ?
            GROUP BY COALESCE(sector, 'Unknown')
            ORDER BY total_value DESC""",
            (user_id,)
        )

        headers = ["Sector", "# Stocks", "Market Value", "Unrealized P&L", "Allocation %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1)

        # Calculate total for percentage
        total_cursor = self.conn.execute(
            "SELECT SUM(market_value) FROM stock_holdings WHERE user_id = ?",
            (user_id,)
        )
        total_value = total_cursor.fetchone()[0] or 1

        for row_idx, row_data in enumerate(cursor.fetchall(), 2):
            sector, count, value, pnl = row_data
            allocation_pct = (value / total_value * 100) if value else 0

            ws.cell(row=row_idx, column=1, value=sector)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=value)
            ws.cell(row=row_idx, column=4, value=pnl)
            ws.cell(row=row_idx, column=5, value=f"{allocation_pct:.2f}%")

    def _add_transactions_sheet(self, wb: Workbook, user_id: int, financial_year: str):
        """Add transactions history sheet."""
        ws = wb.create_sheet("Transactions")

        cursor = self.conn.execute(
            """SELECT
                symbol, isin, company_name, quantity,
                buy_date, sell_date, holding_period_days,
                buy_price, sell_price, buy_value, sell_value,
                taxable_profit, gain_type, quarter
            FROM stock_capital_gains_detail
            WHERE user_id = ? AND financial_year = ?
            ORDER BY sell_date DESC""",
            (user_id, financial_year)
        )

        headers = [
            "Symbol", "ISIN", "Company", "Qty",
            "Buy Date", "Sell Date", "Holding Days",
            "Buy Price", "Sell Price", "Buy Value", "Sell Value",
            "Taxable Profit", "Type", "Quarter"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1)

        for row_idx, row_data in enumerate(cursor.fetchall(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _add_cg_fy_sheet(self, wb: Workbook, user_id: int, financial_year: str):
        """Add FY capital gains summary sheet for ITR."""
        ws = wb.create_sheet("Capital Gains - FY")

        # Get exemption limit from config
        exemption_limit = Decimal(str(self.cg_rules.get("ltcg_exemption_limit", 125000)))

        cursor = self.conn.execute(
            """SELECT
                gain_type,
                COUNT(*) as txn_count,
                SUM(quantity) as total_qty,
                SUM(sell_value) as total_sell,
                SUM(buy_value) as total_buy,
                SUM(taxable_profit) as total_profit,
                SUM(turnover) as total_turnover
            FROM stock_capital_gains_detail
            WHERE user_id = ? AND financial_year = ?
            GROUP BY gain_type""",
            (user_id, financial_year)
        )

        # Headers
        ws.cell(row=1, column=1, value=f"Capital Gains Summary - FY {financial_year}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        self._apply_header_style(ws, 1)

        headers = ["Type", "Transactions", "Total Qty", "Total Sell", "Total Buy", "Profit/Loss", "Turnover"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3)

        stcg_total = Decimal("0")
        ltcg_total = Decimal("0")

        row_idx = 4
        for row_data in cursor.fetchall():
            gain_type, txn_count, total_qty, total_sell, total_buy, total_profit, total_turnover = row_data

            ws.cell(row=row_idx, column=1, value=gain_type)
            ws.cell(row=row_idx, column=2, value=txn_count)
            ws.cell(row=row_idx, column=3, value=total_qty)
            ws.cell(row=row_idx, column=4, value=total_sell)
            ws.cell(row=row_idx, column=5, value=total_buy)
            ws.cell(row=row_idx, column=6, value=total_profit)
            ws.cell(row=row_idx, column=7, value=total_turnover)

            if gain_type == "STCG":
                stcg_total = Decimal(str(total_profit or 0))
            else:
                ltcg_total = Decimal(str(total_profit or 0))

            row_idx += 1

        # ITR Summary
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="ITR Schedule CG Summary")
        self._apply_header_style(ws, row_idx)

        row_idx += 1
        ws.cell(row=row_idx, column=1, value="STCG (Section 111A) - 20%")
        ws.cell(row=row_idx, column=2, value=float(stcg_total))

        row_idx += 1
        ws.cell(row=row_idx, column=1, value="LTCG (Section 112A) - Before Exemption")
        ws.cell(row=row_idx, column=2, value=float(ltcg_total))

        row_idx += 1
        ltcg_exemption = min(exemption_limit, max(Decimal("0"), ltcg_total))
        ws.cell(row=row_idx, column=1, value="LTCG Exemption (u/s 112A - ₹1.25L)")
        ws.cell(row=row_idx, column=2, value=float(ltcg_exemption))

        row_idx += 1
        taxable_ltcg = max(Decimal("0"), ltcg_total - exemption_limit)
        ws.cell(row=row_idx, column=1, value="Taxable LTCG (12.5%)")
        ws.cell(row=row_idx, column=2, value=float(taxable_ltcg))

    def _add_cg_quarterly_sheet(self, wb: Workbook, user_id: int, financial_year: str):
        """Add quarterly capital gains breakdown (for advance tax)."""
        ws = wb.create_sheet("CG - Quarterly")

        # Headers: Stock details as rows, Quarters as columns
        ws.cell(row=1, column=1, value=f"Quarterly Capital Gains - FY {financial_year}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        self._apply_header_style(ws, 1)

        # Column headers
        headers = ["Symbol", "ISIN", "Q1", "Q2", "Q3", "Q4", "Q5", "FY Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3)

        # Get quarterly breakdown by stock
        cursor = self.conn.execute(
            """SELECT
                symbol, isin, quarter,
                SUM(taxable_profit) as profit
            FROM stock_capital_gains_detail
            WHERE user_id = ? AND financial_year = ?
            GROUP BY symbol, isin, quarter
            ORDER BY symbol""",
            (user_id, financial_year)
        )

        # Organize data by stock
        stock_data: Dict[str, Dict[str, Any]] = {}
        for row_data in cursor.fetchall():
            symbol, isin, quarter, profit = row_data
            key = f"{symbol}|{isin}"
            if key not in stock_data:
                stock_data[key] = {"symbol": symbol, "isin": isin, "quarters": {}}
            stock_data[key]["quarters"][quarter] = profit

        # Write data
        row_idx = 4
        for key, data in stock_data.items():
            ws.cell(row=row_idx, column=1, value=data["symbol"])
            ws.cell(row=row_idx, column=2, value=data["isin"])

            fy_total = Decimal("0")
            for col_idx, qtr in enumerate(["Q1", "Q2", "Q3", "Q4", "Q5"], 3):
                val = data["quarters"].get(qtr, 0)
                ws.cell(row=row_idx, column=col_idx, value=val)
                fy_total += Decimal(str(val or 0))

            ws.cell(row=row_idx, column=8, value=float(fy_total))
            row_idx += 1

        # Quarter totals row
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="TOTAL")
        for col_idx, qtr in enumerate(["Q1", "Q2", "Q3", "Q4", "Q5"], 3):
            total = sum(
                Decimal(str(d["quarters"].get(qtr, 0) or 0))
                for d in stock_data.values()
            )
            ws.cell(row=row_idx, column=col_idx, value=float(total))

    def _add_xirr_sheet(self, wb: Workbook, user_id: int):
        """Add XIRR performance sheet."""
        ws = wb.create_sheet("XIRR Performance")

        ws.cell(row=1, column=1, value="XIRR Performance Analysis")
        self._apply_header_style(ws, 1)

        # Overall XIRR
        ws.cell(row=3, column=1, value="Overall Portfolio XIRR")
        if self.result.xirr_overall:
            ws.cell(row=3, column=2, value=f"{self.result.xirr_overall * 100:.2f}%")
        else:
            ws.cell(row=3, column=2, value="N/A")

        # Per-stock XIRR
        ws.cell(row=5, column=1, value="Per-Stock XIRR")
        self._apply_header_style(ws, 5)

        headers = ["Symbol", "XIRR %", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
        self._apply_header_style(ws, 6)

        row_idx = 7
        for symbol, xirr in sorted(
            self.result.xirr_by_stock.items(),
            key=lambda x: x[1] if x[1] else -999,
            reverse=True
        ):
            ws.cell(row=row_idx, column=1, value=symbol)
            if xirr is not None:
                ws.cell(row=row_idx, column=2, value=f"{xirr * 100:.2f}%")
                ws.cell(row=row_idx, column=3, value="Calculated")
            else:
                ws.cell(row=row_idx, column=2, value="N/A")
                ws.cell(row=row_idx, column=3, value="Insufficient data")
            row_idx += 1

    def _add_dividends_sheet(self, wb: Workbook, user_id: int, financial_year: str):
        """Add dividend summary sheet."""
        ws = wb.create_sheet("Dividend Summary")

        cursor = self.conn.execute(
            """SELECT
                symbol, isin,
                COUNT(*) as div_count,
                SUM(gross_amount) as total_gross,
                SUM(tds_amount) as total_tds,
                SUM(net_amount) as total_net
            FROM stock_dividends
            WHERE user_id = ?
            GROUP BY symbol, isin
            ORDER BY total_gross DESC""",
            (user_id,)
        )

        headers = ["Symbol", "ISIN", "# Dividends", "Gross Amount", "TDS", "Net Amount"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1)

        for row_idx, row_data in enumerate(cursor.fetchall(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _apply_header_style(self, ws, row: int):
        """Apply header styling to a row."""
        header_style = self.report_settings.get("header_style", {})
        font = Font(
            bold=header_style.get("bold", True),
            color=header_style.get("font_color", "FFFFFF").replace("#", "")
        )
        fill = PatternFill(
            start_color=header_style.get("bg_color", "4472C4").replace("#", ""),
            end_color=header_style.get("bg_color", "4472C4").replace("#", ""),
            fill_type="solid"
        )

        for cell in ws[row]:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


# =============================================================================
# Main Analyzer (Orchestrator)
# =============================================================================

class StockAnalyzer:
    """
    Main orchestrator for stock statement analysis.

    Coordinates scanning, normalization, ingestion, and reporting.
    """

    def __init__(
        self,
        conn,
        config_path: Optional[Path] = None,
        config: Optional[Dict[str, Any]] = None,
        path_resolver=None
    ):
        """
        Initialize Stock Analyzer.

        Args:
            conn: Database connection
            config_path: Path to config JSON file
            config: Pre-loaded config dict (overrides config_path)
            path_resolver: PathResolver instance for paths
        """
        self.conn = conn
        self.path_resolver = path_resolver

        # Load config
        if config:
            self.config = config
        elif config_path and config_path.exists():
            with open(config_path, encoding="utf-8") as f:
                self.config = json.load(f)
        else:
            # Try default location
            default_config = Path(__file__).parent.parent.parent.parent / "config" / "stock_analyzer_config.json"
            if default_config.exists():
                with open(default_config, encoding="utf-8") as f:
                    self.config = json.load(f)
            else:
                raise FileNotFoundError("Stock analyzer config not found")

        # Initialize components
        self.scanner = StockStatementScanner(self.config, path_resolver)
        self.normalizer = StockFieldNormalizer(self.config)
        self.ingester = StockDBIngester(conn, self.config)

        # Result tracking
        self.result = AnalysisResult()
        self.user_id: Optional[int] = None
        self.financial_year: Optional[str] = None

    def analyze(
        self,
        user_name: str,
        financial_year: str,
        base_path: Optional[Path] = None
    ) -> AnalysisResult:
        """
        Run full analysis pipeline.

        Args:
            user_name: User name
            financial_year: Target FY (e.g., "2025-26")
            base_path: Override base path for scanning

        Returns:
            AnalysisResult with stats and metrics
        """
        self.financial_year = financial_year
        self.result = AnalysisResult()

        # Get or create user
        self.user_id = self._get_or_create_user(user_name)

        # Determine base path
        if base_path is None:
            if self.path_resolver:
                base_path = self.path_resolver.inbox() / "Indian-Stocks"
            else:
                base_path = Path(self.config.get("base_path", "Data/Users/default/inbox/Indian-Stocks"))

        logger.info(f"Starting stock analysis for {user_name}, FY {financial_year}")
        logger.info(f"Scanning: {base_path}")

        # 1. Scan for files
        scanned_files = self.scanner.scan(
            base_path,
            recursive=self.config.get("processing", {}).get("scan_recursive", True),
            include_archive=self.config.get("processing", {}).get("include_archive", False)
        )

        self.result.files_scanned = len(scanned_files)

        if not scanned_files:
            self.result.warnings.append(f"No statement files found in {base_path}")
            return self.result

        # 2. Process each file
        for scanned in scanned_files:
            try:
                self._process_file(scanned)
            except Exception as e:
                self.result.errors.append(f"Error processing {scanned.path.name}: {e}")
                logger.error(f"Error processing {scanned.path}: {e}")

        # 3. Calculate aggregates
        self._calculate_aggregates()

        # 4. Calculate XIRR if enabled
        if self.config.get("processing", {}).get("calculate_xirr", True):
            self._calculate_xirr()

        self.result.success = len(self.result.errors) == 0
        return self.result

    def _get_or_create_user(self, user_name: str) -> int:
        """Get or create user ID."""
        cursor = self.conn.execute(
            "SELECT id FROM users WHERE name = ?",
            (user_name,)
        )
        row = cursor.fetchone()
        if row:
            return row[0]

        # Create user
        cursor = self.conn.execute(
            "INSERT INTO users (name, pan_encrypted, pan_salt) VALUES (?, ?, ?)",
            (user_name, b"", b"")
        )
        self.conn.commit()
        return cursor.lastrowid

    def _process_file(self, scanned: ScannedFile):
        """Process a single scanned file."""
        logger.info(f"Processing: {scanned.path.name} [{scanned.broker.value}:{scanned.statement_type.value}]")

        # Get sheet config
        stmt_config = self.config.get("statement_types", {}).get(
            scanned.statement_type.value, {}
        ).get(scanned.broker.value, {})

        file_format = stmt_config.get("file_format", "auto")
        header_row = stmt_config.get("header_row", 0)

        # Read file based on format
        suffix = scanned.path.suffix.lower()
        df = None

        try:
            if suffix == ".csv" or file_format == "csv_or_tsv":
                # Try to detect CSV vs TSV
                with open(scanned.path, 'r', encoding='utf-8', errors='ignore') as f:
                    first_line = f.readline()
                    if '\t' in first_line:
                        df = pd.read_csv(scanned.path, sep='\t', header=header_row, encoding='utf-8', on_bad_lines='skip')
                    else:
                        df = pd.read_csv(scanned.path, header=header_row, encoding='utf-8', on_bad_lines='skip')
            elif suffix in [".xls", ".xlsx"]:
                # Check if it's actually a TSV masquerading as Excel
                try:
                    with open(scanned.path, 'rb') as f:
                        magic = f.read(8)
                    if magic.startswith(b'PK') or magic[:4] == b'\xd0\xcf\x11\xe0':
                        # Real Excel file
                        sheets = stmt_config.get("sheets", ["Sheet1"])
                        for sheet in sheets:
                            try:
                                df = pd.read_excel(scanned.path, sheet_name=sheet, header=header_row)
                                if not df.empty and len(df.columns) > 1:
                                    break
                            except Exception:
                                continue
                        if df is None or df.empty:
                            df = pd.read_excel(scanned.path, header=header_row)
                    else:
                        # TSV with wrong extension
                        df = pd.read_csv(scanned.path, sep='\t', header=header_row, encoding='utf-8', on_bad_lines='skip')
                except Exception as e:
                    logger.warning(f"Format detection failed, trying Excel: {e}")
                    df = pd.read_excel(scanned.path, header=header_row)
            else:
                self.result.warnings.append(f"Unsupported file format: {scanned.path.name}")
                return
        except Exception as e:
            self.result.errors.append(f"Failed to read {scanned.path.name}: {e}")
            return

        if df is None or df.empty:
            self.result.warnings.append(f"Empty file: {scanned.path.name}")
            return

        # Clean up column names (remove unnamed columns, strip whitespace)
        df.columns = [str(c).strip() if not str(c).startswith('Unnamed') else c for c in df.columns]

        # Normalize and ingest based on statement type
        if scanned.statement_type == StatementType.HOLDINGS:
            holdings = self.normalizer.normalize_holdings(
                df, scanned.broker, str(scanned.path)
            )

            if holdings:
                as_of_date = scanned.modified_date.date() if scanned.modified_date else date.today()
                inserted, skipped = self.ingester.ingest_holdings(
                    holdings, self.user_id, as_of_date
                )
                self.result.holdings_processed += inserted
                self.result.duplicates_skipped += skipped

        elif scanned.statement_type == StatementType.TRANSACTIONS:
            transactions = self.normalizer.normalize_transactions(
                df, scanned.broker, str(scanned.path), self.financial_year
            )

            if transactions:
                inserted, skipped = self.ingester.ingest_transactions(
                    transactions, self.user_id
                )
                self.result.transactions_processed += inserted
                self.result.duplicates_skipped += skipped

        else:
            self.result.warnings.append(
                f"Unknown statement type for {scanned.path.name}, skipping"
            )

    def _calculate_aggregates(self):
        """Calculate aggregate metrics from database."""
        # Holdings totals
        cursor = self.conn.execute(
            """SELECT
                SUM(market_value), SUM(total_cost_basis), SUM(unrealized_pnl)
            FROM stock_holdings WHERE user_id = ?""",
            (self.user_id,)
        )
        row = cursor.fetchone()
        if row:
            self.result.total_market_value = Decimal(str(row[0] or 0))
            self.result.total_cost_basis = Decimal(str(row[1] or 0))
            self.result.total_unrealized_pnl = Decimal(str(row[2] or 0))

        # Capital gains totals
        cursor = self.conn.execute(
            """SELECT gain_type, SUM(taxable_profit)
            FROM stock_capital_gains_detail
            WHERE user_id = ? AND financial_year = ?
            GROUP BY gain_type""",
            (self.user_id, self.financial_year)
        )

        for row in cursor.fetchall():
            if row[0] == "STCG":
                self.result.total_stcg = Decimal(str(row[1] or 0))
            elif row[0] == "LTCG":
                self.result.total_ltcg = Decimal(str(row[1] or 0))

        # LTCG exemption
        exemption_limit = Decimal(str(
            self.config.get("capital_gains_rules", {}).get("ltcg_exemption_limit", 125000)
        ))
        self.result.ltcg_exemption_used = min(
            exemption_limit,
            max(Decimal("0"), self.result.total_ltcg)
        )

        # By broker breakdown
        cursor = self.conn.execute(
            """SELECT b.broker_code, SUM(h.market_value), SUM(h.unrealized_pnl)
            FROM stock_holdings h
            LEFT JOIN stock_brokers b ON h.broker_id = b.id
            WHERE h.user_id = ?
            GROUP BY b.broker_code""",
            (self.user_id,)
        )
        for row in cursor.fetchall():
            self.result.by_broker[row[0]] = {
                "market_value": Decimal(str(row[1] or 0)),
                "unrealized_pnl": Decimal(str(row[2] or 0))
            }

        # By quarter breakdown
        cursor = self.conn.execute(
            """SELECT quarter, SUM(taxable_profit)
            FROM stock_capital_gains_detail
            WHERE user_id = ? AND financial_year = ?
            GROUP BY quarter""",
            (self.user_id, self.financial_year)
        )
        for row in cursor.fetchall():
            self.result.by_quarter[row[0]] = {
                "profit": Decimal(str(row[1] or 0))
            }

    def _calculate_xirr(self):
        """Calculate XIRR for portfolio and individual stocks."""
        # Get all transactions and current holdings
        cursor = self.conn.execute(
            """SELECT symbol, buy_date, sell_date, buy_value, sell_value
            FROM stock_capital_gains_detail
            WHERE user_id = ?
            ORDER BY symbol, buy_date""",
            (self.user_id,)
        )

        # Group transactions by symbol
        by_symbol: Dict[str, List[NormalizedTransaction]] = {}
        all_cashflows: List[Tuple[date, Decimal]] = []

        for row in cursor.fetchall():
            symbol, buy_date_str, sell_date_str, buy_val, sell_val = row

            buy_date = datetime.strptime(buy_date_str, "%Y-%m-%d").date() if buy_date_str else None
            sell_date = datetime.strptime(sell_date_str, "%Y-%m-%d").date() if sell_date_str else None

            txn = NormalizedTransaction(
                symbol=symbol,
                isin="",
                buy_date=buy_date,
                sell_date=sell_date,
                buy_value=Decimal(str(buy_val or 0)),
                sell_value=Decimal(str(sell_val or 0))
            )

            if symbol not in by_symbol:
                by_symbol[symbol] = []
            by_symbol[symbol].append(txn)

            # Add to overall cashflows
            if buy_date and buy_val:
                all_cashflows.append((buy_date, -Decimal(str(buy_val))))
            if sell_date and sell_val:
                all_cashflows.append((sell_date, Decimal(str(sell_val))))

        # Get current holdings for unrealized portion
        cursor = self.conn.execute(
            """SELECT symbol, market_value, as_of_date
            FROM stock_holdings WHERE user_id = ?""",
            (self.user_id,)
        )

        for row in cursor.fetchall():
            symbol, market_val, as_of_str = row
            as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date() if as_of_str else date.today()

            if market_val and market_val > 0:
                all_cashflows.append((as_of, Decimal(str(market_val))))

        # Calculate overall XIRR
        self.result.xirr_overall = XIRRCalculator.calculate(all_cashflows)

        # Calculate per-stock XIRR
        for symbol, txns in by_symbol.items():
            # Get current holding for this stock
            cursor = self.conn.execute(
                """SELECT market_value, as_of_date
                FROM stock_holdings WHERE user_id = ? AND symbol = ?""",
                (self.user_id, symbol)
            )
            holding_row = cursor.fetchone()

            holding = None
            if holding_row and holding_row[0]:
                holding = NormalizedHolding(
                    symbol=symbol,
                    isin="",
                    company_name="",
                    sector=None,
                    quantity_held=0,
                    market_value=Decimal(str(holding_row[0])),
                    as_of_date=datetime.strptime(holding_row[1], "%Y-%m-%d").date() if holding_row[1] else date.today()
                )

            xirr = XIRRCalculator.calculate_for_stock(txns, holding)
            self.result.xirr_by_stock[symbol] = xirr

    def generate_reports(
        self,
        output_path: Optional[Path] = None,
        output_format: str = "xlsx"
    ) -> Path:
        """
        Generate analysis reports.

        Args:
            output_path: Override output path
            output_format: Output format (xlsx, csv)

        Returns:
            Path to generated report
        """
        if output_path is None:
            if self.path_resolver:
                output_path = self.path_resolver.reports() / f"stock_report_{self.financial_year}.xlsx"
            else:
                output_path = Path(f"stock_report_{self.financial_year}.xlsx")

        generator = StockReportGenerator(self.conn, self.config, self.result)
        return generator.generate(output_path, self.user_id, self.financial_year)


# =============================================================================
# Exports
# =============================================================================

__all__ = [
    "StockAnalyzer",
    "StockStatementScanner",
    "StockFieldNormalizer",
    "StockDBIngester",
    "StockReportGenerator",
    "BrokerDetector",
    "XIRRCalculator",
    "AnalysisResult",
    "NormalizedHolding",
    "NormalizedTransaction",
    "ScannedFile",
    "BrokerType",
    "StatementType",
    "GainType",
]
